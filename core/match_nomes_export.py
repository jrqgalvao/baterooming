from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .matching import MATCH_FOUND, check_cancel


_DEFAULT_NAME_COLUMN = 2
_THIN = Border(
    left=Side(style="thin", color="C8D8C8"),
    right=Side(style="thin", color="C8D8C8"),
    top=Side(style="thin", color="C8D8C8"),
    bottom=Side(style="thin", color="C8D8C8"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

_STATUS_STYLES: dict[str, dict] = {
    "Match encontrado": {"fill": "E6F4E7", "font_color": "1E6B22"},
    "Nao encontrado":   {"fill": "FFF3E0", "font_color": "BF360C"},
    "NOME VAZIO":       {"fill": "F4F6F4", "font_color": "7A9A7A"},
}


def write_excel(
    nomes_finais : list[str],
    statuses     : list[str],
    path         : Path,
    scores       : "list[float] | None" = None,
    template_path: "str | Path | None" = None,
    name_rows    : "list[int] | None" = None,
    name_column  : int = _DEFAULT_NAME_COLUMN,
    progress=None,
    should_cancel=None,
) -> None:
    """
    Quando `template_path` e `name_rows` sao fornecidos, salva uma copia da
    Planilha 2 e substitui somente os nomes encontrados na coluna de nomes.

    Sem template, mantem o export legado com colunas:
      nome_final | status | similaridade (%)
    """
    def report(percent: int, message: str) -> None:
        if progress is not None:
            progress(percent, message)

    check_cancel(should_cancel)
    if template_path is not None and name_rows is not None:
        wb = load_workbook(str(template_path))
        try:
            report(20, "Atualizando planilha modelo...")
            ws = wb.active
            for index, (row_idx, nome_final, status) in enumerate(zip(name_rows, nomes_finais, statuses)):
                if index % 128 == 0:
                    check_cancel(should_cancel)
                if status == MATCH_FOUND:
                    ws.cell(row=int(row_idx), column=int(name_column)).value = nome_final
            check_cancel(should_cancel)
            report(85, "Salvando planilha...")
            wb.save(str(path))
            check_cancel(should_cancel)
            report(100, "Planilha concluída.")
        finally:
            wb.close()
        return

    include_score = scores is not None and len(scores) == len(nomes_finais)

    wb = Workbook()
    try:
        report(15, "Preparando planilha...")
        ws = wb.active
        ws.title = "Sheet1"

        n_cols = 3 if include_score else 2
        headers = ["nome_final", "status"]
        if include_score:
            headers.append("similaridade_%")
        ws.append(headers)
        if include_score:
            for index, row in enumerate(zip(nomes_finais, statuses, scores)):
                if index % 128 == 0:
                    check_cancel(should_cancel)
                ws.append(row)
        else:
            for index, row in enumerate(zip(nomes_finais, statuses)):
                if index % 128 == 0:
                    check_cancel(should_cancel)
                ws.append(row)

        # ── Cabeçalho ──
        HDR_FILL  = _fill("141E14")
        HDR_FONT  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        HDR_ALIGN = Alignment(horizontal="center", vertical="center")

        for col in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col)
            c.fill      = HDR_FILL
            c.font      = HDR_FONT
            c.alignment = HDR_ALIGN
            c.border    = _THIN
        ws.row_dimensions[1].height = 26

        # ── Dados ──
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 128 == 0:
                check_cancel(should_cancel)
            status = str(ws.cell(row_idx, 2).value or "")
            style  = _STATUS_STYLES.get(status, {"fill": "FFFFFF", "font_color": "162016"})

            for col in range(1, n_cols + 1):
                c = ws.cell(row_idx, col)
                is_score_col = include_score and col == 3
                c.font = Font(
                    name       = "Segoe UI",
                    color      = style["font_color"],
                    size       = 10,
                    bold       = (col == 2),
                    italic     = is_score_col,
                )
                c.fill      = _fill(style["fill"])
                c.alignment = Alignment(
                    horizontal = "left"   if col == 1 else "center",
                    vertical   = "center",
                )
                c.border = _THIN
            ws.row_dimensions[row_idx].height = 18

        # ── Larguras e freeze ──
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 22
        if include_score:
            ws.column_dimensions["C"].width = 18
        ws.freeze_panes = "A2"

        # ── Rodapé ──
        foot_row = ws.max_row + 2
        last_col_letter = "C" if include_score else "B"
        ws.merge_cells(f"A{foot_row}:{last_col_letter}{foot_row}")
        f = ws.cell(
            foot_row, 1,
            f"{{COMPANY_NAME}} © 2026 — Match de Nomes v3 · Gerado em "
            f"{datetime.now().strftime('%d/%m/%Y %H:%M')}"
        )
        f.font      = Font(italic=True, color="9AAA9A", size=8)
        f.alignment = Alignment(horizontal="right")

        check_cancel(should_cancel)
        report(85, "Salvando planilha...")
        wb.save(str(path))
        check_cancel(should_cancel)
        report(100, "Planilha concluída.")
    finally:
        wb.close()

