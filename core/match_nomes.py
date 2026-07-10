"""
match_nomes.py - Match de Nomes v3
Lógica de fuzzy matching com pipeline multi-etapa.

Pipeline de matching (em ordem de prioridade):
  1. Exato normalizado          — score 100, custo zero
  2. Token set ratio (rapidfuzz) — lida com palavras fora de ordem
  3. Token sort ratio            — lida com inversão de nome/sobrenome
  4. Partial ratio               — lida com nomes abreviados/incompletos
  5. Weighted best               — combinação ponderada dos scores acima

Sem dependências de UI.

Dependências:
    pip install rapidfuzz openpyxl
"""

from openpyxl import load_workbook

from .matching import (
    EMPTY_NAME,
    MATCH_FOUND,
    NOT_FOUND,
    ProcessingCancelled,
    atribuir_melhores_matches,
    check_cancel,
    normalizar,
    tokens_significativos,
)


_NOME_HEADER_ALIASES = {"nome", "name", "hospede", "guest", "cliente", "passageiro"}
_DEFAULT_NAME_COLUMN = 2


def _header_slug(value) -> str:
    return normalizar(str(value or "")).replace(" ", "").replace("_", "")


def _find_name_header(ws) -> "tuple[int, int] | None":
    max_scan_rows = min(5, ws.max_row or 0)
    max_scan_cols = ws.max_column or 0
    for row_idx in range(1, max_scan_rows + 1):
        for col_idx in range(1, max_scan_cols + 1):
            if _header_slug(ws.cell(row=row_idx, column=col_idx).value) in _NOME_HEADER_ALIASES:
                return col_idx, row_idx
    return None


def _detect_name_layout(ws) -> tuple[int, int]:
    header = _find_name_header(ws)
    if header is not None:
        name_col, header_row = header
        return name_col, header_row + 1

    if (ws.max_column or 0) >= 4:
        name_col = _DEFAULT_NAME_COLUMN
    else:
        name_col = 1

    return name_col, 1


def _read_nomes(path: str, should_cancel=None) -> tuple[list[object], list[int], int]:
    check_cancel(should_cancel)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws.max_row == 0:
            return [], [], _DEFAULT_NAME_COLUMN

        name_col, start_row = _detect_name_layout(ws)
        max_row = ws.max_row or 0
        nomes = []
        row_numbers = []
        for row_idx, row in enumerate(
            ws.iter_rows(
                min_row=start_row,
                max_row=max_row,
                min_col=name_col,
                max_col=name_col,
                values_only=True,
            ),
            start=start_row,
        ):
            if row_idx % 256 == 0:
                check_cancel(should_cancel)
            nomes.append(row[0] if row else None)
            row_numbers.append(row_idx)
        return nomes, row_numbers, name_col
    finally:
        wb.close()


def _texto_planilha(value) -> str:
    return str(value).strip()


def _nome_limpo(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text


# ── Scoring multi-etapa ────────────────────────────────────────

def executar_match(
    arquivo_a: str,
    arquivo_b: str,
    threshold: int,
    progress=None,
    should_cancel=None,
) -> dict:
    """
    Executa o fuzzy match entre Lista A (referência) e Lista B (a corrigir).

    Regras de negócio:
      - A saída tem exatamente len(Lista B) linhas, na mesma ordem.
      - Linhas sem match suficiente mantêm o nome original e recebem
        status "Nao encontrado".
      - Linhas vazias recebem status "NOME VAZIO".

    Retorna:
        {
            "ok": True,
            "nomes_finais": [...],
            "statuses":     [...],
            "scores":       [...],   # float 0-100, para debug/auditoria
            "kpis": {"total": N, "match": N, "nomatch": N, "empty": N}
        }
        ou {"ok": False, "erro": "..."}
    """
    def report(percent: int, message: str) -> None:
        if progress is not None:
            progress(percent, message)

    try:
        report(10, "Lendo lista de referência...")
        nomes_a_raw, _, _ = _read_nomes(arquivo_a, should_cancel=should_cancel)
        report(25, "Lendo lista para padronizar...")
        nomes_b_raw, name_rows, name_column = _read_nomes(arquivo_b, should_cancel=should_cancel)

        nomes_a = [_texto_planilha(value) for value in nomes_a_raw]
        nomes_b_texto = [_texto_planilha(value) for value in nomes_b_raw]

        # Remove linhas totalmente vazias da lista de referência
        lista_a_original    = [nome for nome in nomes_a if nome not in {"", "nan", "None"}]
        lista_a_normalizada = [normalizar(n) for n in lista_a_original]
        lista_a_tokens      = [set(tokens_significativos(n)) for n in lista_a_normalizada]
        thr_float = float(threshold)
        nomes_b = [_nome_limpo(nome_b_raw) for nome_b_raw in nomes_b_texto]
        nomes_b_normalizados = [normalizar(nome_b_clean) for nome_b_clean in nomes_b]

        report(40, "Comparando nomes...")

        def matching_progress(fraction: float) -> None:
            report(40 + round(45 * fraction), "Comparando nomes...")

        nomes_finais, statuses, scores = atribuir_melhores_matches(
            nomes_b,
            nomes_b_normalizados,
            lista_a_original,
            lista_a_normalizada,
            lista_a_tokens,
            thr_float,
            progress=matching_progress,
            should_cancel=should_cancel,
        )

        total = len(nomes_b)
        match_count = statuses.count(MATCH_FOUND)
        nomatch_count = statuses.count(NOT_FOUND)
        empty_count = statuses.count(EMPTY_NAME)

        kpis = {
            "total":   total,
            "match":   match_count,
            "nomatch": nomatch_count,
            "empty":   empty_count,
        }

        check_cancel(should_cancel)
        report(90, "Consolidando resultado...")
        return {
            "ok":           True,
            "nomes_finais": nomes_finais,
            "statuses":     statuses,
            "scores":       scores,
            "kpis":         kpis,
            "template_path": str(arquivo_b),
            "name_rows":     name_rows,
            "name_column":   name_column if name_rows else _DEFAULT_NAME_COLUMN,
        }

    except ProcessingCancelled:
        raise
    except Exception as exc:
        return {"ok": False, "erro": str(exc).strip() or "Não foi possível processar as planilhas."}

