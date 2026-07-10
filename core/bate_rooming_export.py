from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .bate_rooming import Status, _is_room_ok, calc_kpis
from .matching import check_cancel


# ─── EXPORTAÇÃO EXCEL ─────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _font(*, bold=False, italic=False, color="162016", size=10) -> Font:
    return Font(bold=bold, italic=italic, color=color.lstrip("#"), size=size)


_THIN = Border(
    left=Side(style="thin", color="C8D8C8"),
    right=Side(style="thin", color="C8D8C8"),
    top=Side(style="thin", color="C8D8C8"),
    bottom=Side(style="thin", color="C8D8C8"),
)

ALIGN_LEFT        = Alignment(vertical="center", horizontal="left")
ALIGN_LEFT_WRAP   = Alignment(vertical="center", horizontal="left", wrap_text=True)
ALIGN_CENTER      = Alignment(vertical="center", horizontal="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Fills de linha por categoria (para filtrar por cor no Excel)
FILL_ROW_OK     = _fill("F7FBF7")   # verde
FILL_ROW_DIV    = _fill("FFFDF4")   # amarelo
FILL_ROW_NOMAP  = _fill("F8F6FC")   # roxo
FILL_ROW_REP    = _fill("FDF7FA")   # rosa
FILL_ROW_ABS    = _fill("FFFBF5")   # laranja

def _row_fill(r: dict) -> "PatternFill":
    """Retorna o fill correto para a linha inteira baseado na categoria dominante."""
    if r.get("is_dup") or r.get("s_quarto") == "REPETIDO":
        return FILL_ROW_REP
    if r.get("no_match"):
        return FILL_ROW_NOMAP
    statuses = [r.get("s_quarto"), r.get("s_ci"), r.get("s_co")]
    if any(s == Status.DATA_AUSENTE for s in statuses):
        return FILL_ROW_ABS
    if _is_room_ok(r.get("s_quarto")) and r.get("s_ci") == Status.OK and r.get("s_co") == Status.OK:
        return FILL_ROW_OK
    return FILL_ROW_DIV

STYLE_CACHE = {
    "ok":         (_fill("E6F4E7"), _font(color="1E6B22")),
    "nomap":      (_fill("EDE7F6"), _font(bold=True, color="4527A0")),
    "absent":     (_fill("FFF3E0"), _font(bold=True, color="BF360C")),
    "div":        (_fill("FFEBEE"), _font(bold=True, color="B71C1C")),
    "averificar": (_fill("FFF8E1"), _font(bold=True, color="F57F17")),
    "repetido":   (_fill("FCE4EC"), _font(bold=True, color="880E4F")),
    "text":       _font(color="162016"),
    "title":      _font(bold=True, color="FFFFFF", size=11),
    "sum_title":  _font(bold=True, color="FFFFFF", size=13),
    "sum_sub":    _font(italic=True, color="6B8F6B", size=9),
}

_XL_HEADERS = [
    "NOME", "FONTE",
    "QUARTO (PLAN. 1)", "QUARTO (PLAN. 2)",
    "CHECK-IN (PLAN. 1)", "CHECK-IN (PLAN. 2)",
    "CHECK-OUT (PLAN. 1)", "CHECK-OUT (PLAN. 2)",
    "STATUS GERAL", "STATUS QUARTO", "STATUS CHECK-IN", "STATUS CHECK-OUT",
]
_XL_WIDTHS = [38, 10, 18, 18, 20, 20, 20, 20, 28, 22, 22, 22]
_XL_ROOM_COLUMNS = {2, 3, 9}
_XL_STATUS_START_INDEX = 8

_LOG_HEADERS = [
    "ORDEM", "FONTE", "NOME NO RESULTADO", "NOME PLANILHA 1", "NOME PLANILHA 2",
    "DECISAO", "ETAPA", "SCORE", "THRESHOLD", "CANDIDATOS CONSIDERADOS",
    "CANDIDATOS ELEGIVEIS", "CANDIDATOS BLOQUEADOS", "TOP CANDIDATOS",
    "CONFLITOS", "MOTIVO",
]
_LOG_WIDTHS = [8, 12, 32, 32, 32, 18, 30, 10, 12, 18, 18, 18, 70, 70, 70]


def _hide_room_columns(results: list) -> bool:
    return any(r.get("room_check_ignored") for r in results)


def _visible_excel_columns(hide_room: bool = False) -> list[int]:
    if not hide_room:
        return list(range(len(_XL_HEADERS)))
    return [idx for idx in range(len(_XL_HEADERS)) if idx not in _XL_ROOM_COLUMNS]


def _status_style(v: str):
    if v == Status.OK:
        return STYLE_CACHE["ok"]
    if v == Status.IGNORADO:
        return STYLE_CACHE["averificar"]
    if v == Status.SEM_CORRESPONDENCIA:
        return STYLE_CACHE["nomap"]
    if v == Status.DATA_AUSENTE:
        return STYLE_CACHE["absent"]
    if v == Status.A_VERIFICAR:
        return STYLE_CACHE["averificar"]
    if v == Status.REPETIDO:
        return STYLE_CACHE["repetido"]
    return STYLE_CACHE["div"]


def _set_header_cell(ws, row, col, text, bg="1B5E20", fg="FFFFFF", sz=10):
    c = ws.cell(row, col, text)
    c.font      = Font(bold=True, color=fg, size=sz)
    c.fill      = _fill(bg)
    c.alignment = ALIGN_CENTER_WRAP
    c.border    = _THIN
    return c


def _all_result_vals(r: dict) -> list:
    """Extrai a lista de valores de uma linha de resultado para uso no Excel."""
    return [
        r["nome"], r["fonte"],
        r["q_sys"], r["q_hotel"],
        r["ci_sys"], r["ci_hotel"],
        r["co_sys"], r["co_hotel"],
        r.get("s_geral", r["s_quarto"]), r["s_quarto"], r["s_ci"], r["s_co"],
    ]


def _result_vals(r: dict, hide_room: bool = False) -> list:
    vals = _all_result_vals(r)
    return [vals[idx] for idx in _visible_excel_columns(hide_room)]


def _apply_data_row(ws, ri, vals, row_fill=None, status_start_col: int = 9):
    if row_fill is None:
        row_fill = _fill("FFFFFF")
    base_font = STYLE_CACHE["text"]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v)
        c.border    = _THIN
        c.alignment = ALIGN_LEFT if ci == 1 else ALIGN_CENTER
        if ci >= status_start_col:
            fill, font = _status_style(str(v))
            c.fill = fill
            c.font = font
        else:
            c.fill = row_fill
            c.font = base_font
    ws.row_dimensions[ri].height = 18


def _apply_filter(ws, header_row: int, last_row: int) -> None:
    if last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{last_row}"


def _build_result_sheet(
    ws,
    results: list,
    title: str = "RESULTADO",
    hide_room: bool = False,
    should_cancel=None,
):
    columns = _visible_excel_columns(hide_room)
    status_start_col = next(
        (i for i, idx in enumerate(columns, 1) if idx >= _XL_STATUS_START_INDEX),
        len(columns) + 1,
    )
    ws.title = title
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    t = ws.cell(1, 1, f"Bate-Rooming — {title} • Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    t.font      = STYLE_CACHE["title"]
    t.fill      = _fill("141E14")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    for ci, idx in enumerate(columns, 1):
        _set_header_cell(ws, 2, ci, _XL_HEADERS[idx])
        ws.column_dimensions[get_column_letter(ci)].width = _XL_WIDTHS[idx]
    ws.row_dimensions[2].height = 30
    for ri, r in enumerate(results, 3):
        if (ri - 3) % 128 == 0:
            check_cancel(should_cancel)
        _apply_data_row(
            ws,
            ri,
            _result_vals(r, hide_room=hide_room),
            row_fill=_row_fill(r),
            status_start_col=status_start_col,
        )
    _apply_filter(ws, 2, len(results) + 2)
    # Linha de totalizador por categoria
    kpi = calc_kpis(results)
    footer_row = len(results) + 4
    ws.merge_cells(f"A{footer_row}:{get_column_letter(len(columns))}{footer_row}")
    summary_text = (
        f"TOTAL: {kpi['total']}  |  "
        f"✓ OK: {kpi['ok']}  |  "
        f"⚠ Divergentes: {kpi['div']}  |  "
        f"? Sem correspondência: {kpi['nomap']}  |  "
        f"🔁 Repetidos: {kpi['dups']}"
    )
    fc = ws.cell(footer_row, 1, summary_text)
    fc.font      = Font(bold=True, color="FFFFFF", size=9)
    fc.fill      = _fill("1B3A1B")
    fc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[footer_row].height = 20


def _build_summary_sheet(ws, results: list, hide_room: bool = False):
    ws.title = "RESUMO"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.merge_cells("A1:C1")
    t = ws.cell(1, 1, "Bate-Rooming — Resumo da Conferência")
    t.font      = STYLE_CACHE["sum_title"]
    t.fill      = _fill("141E14")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:C2")
    d = ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
    d.font      = STYLE_CACHE["sum_sub"]
    d.fill      = _fill("141E14")
    d.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18
    kpi = calc_kpis(results)
    kpis_geral = [
        ("Total de hóspedes",           kpi["total"],   "162016", "F4F6F4", "B8CCB8"),
        ("Verificar manualmente (?)",   kpi["nomap"],   "4527A0", "EDE7F6", "7E57C2"),
        ("Nomes repetidos (REPETIDO)",     kpi["dups"], "880E4F", "FCE4EC", "CE93D8"),
        ("Com correspondência",         kpi["matched"], "162016", "F4F6F4", "B8CCB8"),
        ("Sem divergência ✓",           kpi["ok"],      "1E6B22", "E6F4E7", "1E6B22"),
        ("Com divergência ⚠",           kpi["div"],     "BF360C", "FFF3E0", "BF360C"),
    ]
    kpis_div = []
    if not hide_room:
        kpis_div.append(("Divergência — Quarto", kpi["q_div"], "B71C1C", "FFEBEE", "B71C1C"))
    kpis_div.extend([
        ("Divergência — Check-in",  kpi["ci_div"], "B71C1C", "FFEBEE", "B71C1C"),
        ("Divergência — Check-out", kpi["co_div"], "B71C1C", "FFEBEE", "B71C1C"),
    ])

    def write_section(start_row, section_label, section_color, rows):
        ws.merge_cells(f"A{start_row}:C{start_row}")
        h = ws.cell(start_row, 1, section_label)
        h.font      = Font(bold=True, color="FFFFFF", size=10)
        h.fill      = _fill(section_color)
        h.alignment = ALIGN_CENTER
        ws.row_dimensions[start_row].height = 22
        _set_header_cell(ws, start_row + 1, 1, "MÉTRICA",    section_color)
        _set_header_cell(ws, start_row + 1, 2, "QUANTIDADE", section_color)
        _set_header_cell(ws, start_row + 1, 3, "% DO TOTAL", section_color)
        ws.row_dimensions[start_row + 1].height = 20
        for i, (label, val, fg, bg, brd) in enumerate(rows, start_row + 2):
            row_fill = _fill(bg)
            for ci in range(1, 4):
                cell = ws.cell(i, ci)
                cell.fill = row_fill
                cell.border = _THIN
            lbl_c           = ws.cell(i, 1, label)
            lbl_c.font      = Font(color=fg, size=10)
            lbl_c.alignment = ALIGN_LEFT
            val_c           = ws.cell(i, 2, val)
            val_c.font      = Font(bold=True, color=fg, size=12)
            val_c.alignment = ALIGN_CENTER
            pct             = f"{val/kpi['total']*100:.1f}%" if kpi["total"] else "—"
            pct_c           = ws.cell(i, 3, pct)
            pct_c.font      = Font(color=fg, size=10)
            pct_c.alignment = ALIGN_CENTER
            ws.row_dimensions[i].height = 22
        return start_row + 2 + len(rows)

    next_row = write_section(4, "VISÃO GERAL", "162016", kpis_geral)
    next_row += 1
    next_row = write_section(next_row, "DETALHES DE DIVERGÊNCIA", "B71C1C", kpis_div)
    foot_row = next_row + 1
    ws.merge_cells(f"A{foot_row}:C{foot_row}")
    f = ws.cell(foot_row, 1, "{{COMPANY_NAME}} © 2026 — Bate-Rooming  |  Abas: RESULTADO COMPLETO · DIVERGÊNCIAS · SEM CORRESPONDÊNCIA · REPETIDOS")
    f.font      = Font(italic=True, color="9AAA9A", size=8)
    f.alignment = ALIGN_CENTER


def _build_sheet_header(ws, title: str, bg: str, hide_room: bool = False):
    """Cria a linha de título e configura a aba."""
    columns = _visible_excel_columns(hide_room)
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    t = ws.cell(1, 1, f"{title} • Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    t.font      = STYLE_CACHE["title"]
    t.fill      = _fill(bg)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _build_data_section(
    ws,
    rows: list,
    bg: str,
    start_row: int,
    empty_msg: str,
    hide_room: bool = False,
    should_cancel=None,
) -> int:
    """Adiciona cabeçalho + linhas de dados. Retorna próxima linha disponível."""
    columns = _visible_excel_columns(hide_room)
    status_start_col = next(
        (i for i, idx in enumerate(columns, 1) if idx >= _XL_STATUS_START_INDEX),
        len(columns) + 1,
    )
    cur_row = start_row
    if rows:
        header_row = cur_row
        for ci, idx in enumerate(columns, 1):
            _set_header_cell(ws, cur_row, ci, _XL_HEADERS[idx], bg=bg)
            ws.column_dimensions[get_column_letter(ci)].width = _XL_WIDTHS[idx]
        ws.row_dimensions[cur_row].height = 30
        cur_row += 1
        for row_index, r in enumerate(rows):
            if row_index % 128 == 0:
                check_cancel(should_cancel)
            _apply_data_row(
                ws,
                cur_row,
                _result_vals(r, hide_room=hide_room),
                row_fill=_row_fill(r),
                status_start_col=status_start_col,
            )
            cur_row += 1
        _apply_filter(ws, header_row, cur_row - 1)
    else:
        ws.merge_cells(f"A{cur_row}:{get_column_letter(len(columns))}{cur_row}")
        c           = ws.cell(cur_row, 1, empty_msg)
        c.font      = Font(bold=True, color="1E6B22", size=11)
        c.fill      = _fill("E6F4E7")
        c.alignment = ALIGN_CENTER
        ws.row_dimensions[cur_row].height = 30
        cur_row += 1
    return cur_row


def _categorize_export_rows(results: list) -> tuple[list, list, list]:
    divergent = []
    nomap = []
    repeated = []
    for r in results:
        is_repeated = r.get("is_dup") or r.get("s_quarto") == "REPETIDO"
        if is_repeated:
            repeated.append(r)
        elif r["no_match"]:
            nomap.append(r)
        elif not (_is_room_ok(r["s_quarto"]) and r["s_ci"] == Status.OK and r["s_co"] == Status.OK):
            divergent.append(r)
    return divergent, nomap, repeated


def _build_divergences_sheet(ws, divergent: list, hide_room: bool = False, should_cancel=None):
    ws.title = "DIVERGÊNCIAS"
    _build_sheet_header(ws,
        f"DIVERGÊNCIAS — {len(divergent)} registro(s) com dados divergentes entre as planilhas",
        "B71C1C", hide_room=hide_room)
    _build_data_section(
        ws,
        divergent,
        "B71C1C",
        2,
        "✓ Nenhuma divergência encontrada!",
        hide_room=hide_room,
        should_cancel=should_cancel,
    )
    ws.freeze_panes = "A3"


def _build_nomap_sheet(ws, nomap: list, hide_room: bool = False, should_cancel=None):
    ws.title = "SEM CORRESPONDÊNCIA"
    _build_sheet_header(ws,
        f"SEM CORRESPONDÊNCIA — {len(nomap)} registro(s) sem par entre as planilhas",
        "4527A0", hide_room=hide_room)
    _build_data_section(
        ws,
        nomap,
        "4527A0",
        2,
        "✓ Nenhum registro sem correspondência!",
        hide_room=hide_room,
        should_cancel=should_cancel,
    )
    ws.freeze_panes = "A3"


def _build_repeated_sheet(ws, repeated: list, hide_room: bool = False, should_cancel=None):
    ws.title = "REPETIDOS"
    _build_sheet_header(ws,
        f"REPETIDOS — {len(repeated)} registro(s) com nome duplicado ou placeholder",
        "880E4F", hide_room=hide_room)
    _build_data_section(
        ws,
        repeated,
        "880E4F",
        2,
        "✓ Nenhum nome repetido encontrado!",
        hide_room=hide_room,
        should_cancel=should_cancel,
    )
    ws.freeze_panes = "A3"



def _format_log_score(value) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):.1f}"
    except (TypeError, ValueError):
        return str(value)


def _audit_for_export(r: dict) -> dict:
    audit = r.get("match_audit")
    if isinstance(audit, dict):
        return audit
    if r.get("is_dup") or r.get("s_quarto") == "REPETIDO":
        decision = "REPETIDO"
    elif r.get("no_match"):
        decision = "SEM_MATCH"
    else:
        decision = "MATCH"
    return {
        "decision": decision,
        "stage": "SEM_LOG",
        "score": None,
        "threshold": "",
        "candidates_considered": 0,
        "candidates_eligible": 0,
        "candidates_blocked": 0,
        "top_candidates": [],
        "conflicts": [],
        "ambiguous": False,
        "reason": "Resultado sem detalhes de matching disponíveis.",
    }


def _log_candidates_text(audit: dict) -> str:
    values = []
    for item in audit.get("top_candidates", []):
        name = str(item.get("nome", "")).strip()
        if not name:
            continue
        score = _format_log_score(item.get("score"))
        stage = str(item.get("stage", "")).strip()
        label = name
        if score:
            label += f" ({score})"
        if stage:
            label += f" [{stage}]"
        values.append(label)
    return "; ".join(values)


def _log_conflicts_text(audit: dict) -> str:
    values = []
    for item in audit.get("conflicts", []):
        reference = str(item.get("nome_referencia", "")).strip()
        winner = str(item.get("winner_nome", "")).strip()
        score = _format_log_score(item.get("score"))
        winner_score = _format_log_score(item.get("winner_score"))
        if reference and winner:
            values.append(f"{reference} ({score}) -> {winner} ({winner_score})")
        elif reference:
            values.append(f"{reference} ({score})")
    return "; ".join(values)


def _log_row_fill(audit: dict) -> PatternFill:
    decision = audit.get("decision")
    if decision == "REPETIDO":
        return FILL_ROW_REP
    if audit.get("ambiguous") or audit.get("candidates_blocked") or audit.get("conflicts"):
        return FILL_ROW_ABS
    if decision == "SEM_MATCH":
        return FILL_ROW_NOMAP
    return FILL_ROW_OK


def _build_log_sheet(ws, results: list, should_cancel=None):
    ws.title = "LOG"
    ws.sheet_view.showGridLines = False
    last_col = get_column_letter(len(_LOG_HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    title = ws.cell(1, 1, "Bate-Rooming - LOG DE MATCHING E RASTREABILIDADE")
    title.font = STYLE_CACHE["title"]
    title.fill = _fill("141E14")
    title.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 26

    for ci, header in enumerate(_LOG_HEADERS, 1):
        _set_header_cell(ws, 2, ci, header, bg="455A64")
        ws.column_dimensions[get_column_letter(ci)].width = _LOG_WIDTHS[ci - 1]
    ws.row_dimensions[2].height = 34

    for row_index, result in enumerate(results, 1):
        if (row_index - 1) % 128 == 0:
            check_cancel(should_cancel)
        audit = _audit_for_export(result)
        values = [
            row_index,
            result.get("fonte", ""),
            result.get("nome", ""),
            audit.get("nome_sistema", ""),
            audit.get("nome_hotel", ""),
            audit.get("decision", ""),
            audit.get("stage", ""),
            audit.get("score"),
            audit.get("threshold", ""),
            audit.get("candidates_considered", 0),
            audit.get("candidates_eligible", 0),
            audit.get("candidates_blocked", 0),
            _log_candidates_text(audit),
            _log_conflicts_text(audit),
            audit.get("reason", ""),
        ]
        row = row_index + 2
        fill = _log_row_fill(audit)
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = _THIN
            cell.fill = fill
            cell.font = Font(bold=col in (6, 7), color="162016", size=10)
            cell.alignment = (
                ALIGN_LEFT_WRAP
                if col in (3, 4, 5, 7, 13, 14, 15)
                else ALIGN_CENTER
            )
        ws.row_dimensions[row].height = 42 if audit.get("ambiguous") else 22

    if not results:
        ws.merge_cells(f"A3:{last_col}3")
        empty = ws.cell(3, 1, "Nenhum resultado processado.")
        empty.font = Font(bold=True, color="1E6B22", size=11)
        empty.fill = FILL_ROW_OK
        empty.alignment = ALIGN_CENTER
        ws.row_dimensions[3].height = 28
    _apply_filter(ws, 2, len(results) + 2)
    ws.freeze_panes = "A3"


def write_excel(results: list, path: Path, progress=None, should_cancel=None):
    def report(percent: int, message: str) -> None:
        if progress is not None:
            progress(percent, message)

    check_cancel(should_cancel)
    hide_room = _hide_room_columns(results)
    divergent, nomap, repeated = _categorize_export_rows(results)
    wb         = openpyxl.Workbook()
    try:
        report(8, "Gerando resumo do relatório...")
        ws_summary = wb.active
        _build_summary_sheet(ws_summary, results, hide_room=hide_room)
        check_cancel(should_cancel)
        report(20, "Gerando resultado completo...")
        ws_result  = wb.create_sheet()
        _build_result_sheet(
            ws_result,
            results,
            "RESULTADO COMPLETO",
            hide_room=hide_room,
            should_cancel=should_cancel,
        )
        check_cancel(should_cancel)
        report(55, "Gerando divergências...")
        ws_divs    = wb.create_sheet()
        _build_divergences_sheet(ws_divs, divergent, hide_room=hide_room, should_cancel=should_cancel)
        check_cancel(should_cancel)
        report(68, "Gerando registros sem correspondência...")
        ws_nomap   = wb.create_sheet()
        _build_nomap_sheet(ws_nomap, nomap, hide_room=hide_room, should_cancel=should_cancel)
        check_cancel(should_cancel)
        report(80, "Gerando registros repetidos...")
        ws_rep     = wb.create_sheet()
        _build_repeated_sheet(ws_rep, repeated, hide_room=hide_room, should_cancel=should_cancel)
        check_cancel(should_cancel)
        report(86, "Gerando log de matching...")
        ws_log = wb.create_sheet()
        _build_log_sheet(ws_log, results, should_cancel=should_cancel)
        check_cancel(should_cancel)
        wb.active  = ws_summary
        report(92, "Salvando relatório...")
        wb.save(path)
        check_cancel(should_cancel)
        report(100, "Relatório concluído.")
    finally:
        wb.close()


# ─── FUNÇÃO DE ALTO NÍVEL: processar dois arquivos ────────────

