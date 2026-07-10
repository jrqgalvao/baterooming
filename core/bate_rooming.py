"""
bate_rooming.py — Bate-Rooming v5
Módulo de lógica pura: leitura e comparação.
Sem dependências de UI (sem Tkinter).
"""

import os
import re
import tempfile
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

from .matching import check_cancel, normalizar


_RE_SPACES = re.compile(r"\s+")
_RE_SLUG_SEPARATORS = re.compile(r"[\s_\-\./]+")
NAME_MATCH_THRESHOLD = 65
CROSS_IDENTITY_MIN_SCORE = 85


def _empty_match_audit(
    *,
    nome_sistema: str = "",
    nome_hotel: str = "",
    decision: str = "SEM_MATCH",
    stage: str = "SEM_MATCH",
    reason: str = "",
) -> dict:
    return {
        "nome_sistema": nome_sistema,
        "nome_hotel": nome_hotel,
        "decision": decision,
        "stage": stage,
        "score": None,
        "threshold": float(NAME_MATCH_THRESHOLD),
        "candidates_considered": 0,
        "candidates_eligible": 0,
        "candidates_blocked": 0,
        "top_candidates": [],
        "conflicts": [],
        "ambiguous": False,
        "log_type": decision,
        "reason": reason,
    }


def _build_match_audit(
    hotel_name: str,
    sys_items: list[tuple[str, dict]],
    detail: dict,
    threshold: float,
    hotel_items: list[tuple[str, dict]] | None = None,
) -> dict:
    assignment_index = detail.get("assignment_index", -1)
    matched_system_name = ""
    if isinstance(assignment_index, int) and 0 <= assignment_index < len(sys_items):
        matched_system_name = sys_items[assignment_index][1]["nome"]

    top_candidates = []
    for item in detail.get("top_candidates", []):
        a_idx = item.get("a_idx", -1)
        if not isinstance(a_idx, int) or not 0 <= a_idx < len(sys_items):
            continue
        top_candidates.append({
            "nome": sys_items[a_idx][1]["nome"],
            "score": item.get("score"),
            "stage": item.get("stage", ""),
            "reason": item.get("reason", ""),
        })

    conflicts = []
    for item in detail.get("conflicts", []):
        a_idx = item.get("a_idx", -1)
        winner_idx = item.get("winner_b_idx")
        conflicts.append({
            "nome_referencia": (
                sys_items[a_idx][1]["nome"]
                if isinstance(a_idx, int) and 0 <= a_idx < len(sys_items)
                else ""
            ),
            "score": item.get("score"),
            "winner_score": item.get("winner_score"),
            "winner_nome": (
                hotel_items[winner_idx][1]["nome"]
                if (
                    hotel_items is not None
                    and isinstance(winner_idx, int)
                    and 0 <= winner_idx < len(hotel_items)
                )
                else ""
            ),
            "reason": item.get("reason", ""),
        })

    stage = detail.get("stage", "SEM_MATCH")
    score = detail.get("score")
    matched = assignment_index >= 0
    ambiguous = len(top_candidates) > 1 or bool(conflicts) or stage == "FUZZY_GLOBAL"
    if matched:
        if stage == "EXATO_NORMALIZADO":
            reason = "Match exato após normalização de acentos, caixa, pontuação e espaços."
        elif stage == "MESMA_IDENTIDADE_MAIOR_SCORE":
            reason = (
                "Maior score entre candidatos com o mesmo primeiro e último token significativo; "
                "a prioridade de identidade foi aplicada antes do fuzzy global."
            )
        else:
            reason = (
                f"Match fuzzy acima do threshold conservador de {threshold:.1f}; "
                "identidades diferentes, revisar o LOG."
            )
        if conflicts:
            reason += " Outros candidatos foram bloqueados por conflito 1:1."
        decision = "MATCH"
    elif conflicts:
        reason = "Nenhum match atribuído: os candidatos disponíveis perderam a referência em conflito 1:1."
        decision = "SEM_MATCH"
    elif detail.get("candidates_blocked", 0):
        reason = "Nenhum match atribuído: candidatos bloqueados pela regra conservadora contra falso positivo."
        decision = "SEM_MATCH"
    elif top_candidates:
        reason = f"Nenhum candidato atingiu as regras de match e o threshold de {threshold:.1f}."
        decision = "SEM_MATCH"
    else:
        reason = "Nenhum candidato disponível na planilha de referência."
        decision = "SEM_MATCH"

    if decision == "MATCH" and ambiguous:
        log_type = "MATCH_AMBIGUO"
    elif decision == "SEM_MATCH" and (ambiguous or detail.get("candidates_blocked", 0)):
        log_type = "SEM_MATCH_AMBIGUO"
    else:
        log_type = decision

    return {
        "nome_sistema": matched_system_name,
        "nome_hotel": hotel_name,
        "decision": decision,
        "stage": stage,
        "score": score,
        "threshold": float(threshold),
        "candidates_considered": detail.get("candidates_considered", 0),
        "candidates_eligible": detail.get("candidates_eligible", 0),
        "candidates_blocked": detail.get("candidates_blocked", 0),
        "top_candidates": top_candidates,
        "conflicts": conflicts,
        "ambiguous": ambiguous,
        "log_type": log_type,
        "reason": reason,
    }

# ─── STATUS ───────────────────────────────────────────────────
# StrEnum foi adicionado no Python 3.11; fallback limpo para 3.10.
try:
    from enum import StrEnum
except ImportError:
    from enum import Enum

    class StrEnum(str, Enum):  # type: ignore
        pass


class Status(StrEnum):
    OK = "OK"
    DIVERGENTE = "DIVERGENTE"
    IGNORADO = "IGNORADO"
    CHECK_IN = "CHECK-IN"
    CHECK_OUT = "CHECK-OUT"
    DATA_AUSENTE = "DATA AUSENTE"
    SEM_CORRESPONDENCIA = "SEM CORRESPONDÊNCIA"
    A_VERIFICAR = "A VERIFICAR"
    REPETIDO = "REPETIDO"


# ─── UTILITÁRIOS ──────────────────────────────────────────────
def strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", str(text))
        if unicodedata.category(c) != "Mn"
    )


def normalize_key(text: str) -> str:
    clean = _RE_SPACES.sub(" ", str(text or "").strip())
    return strip_accents(clean).lower()



# Nomes placeholder que devem sempre ser tratados como REPETIDO
PLACEHOLDER_NAMES = {
    normalizar(name)
    for name in (
        "a nomear", "a definir", "tbd", "to be defined",
        "to be confirmed", "tbc", "sem nome", "n/a", "none",
        "anomear", "adefinir", "semhospede", "sem hospede",
    )
}

def display_name(text: str) -> str:
    return " ".join(
        w.capitalize()
        for w in _RE_SPACES.sub(" ", str(text or "").strip()).split()
    )


_FMT_LENGTHS = {
    "%d/%m/%Y": 10, "%d/%m/%y": 8,
    "%Y-%m-%d %H:%M:%S": 19, "%Y-%m-%d": 10,
    "%d-%m-%Y": 10, "%d.%m.%Y": 10, "%m/%d/%Y": 10,
}


def parse_date_raw(value) -> "date | None":
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s or s in ("None", "—", "-", ""):
        return None
    for fmt, length in _FMT_LENGTHS.items():
        try:
            return datetime.strptime(s[:length], fmt).date()
        except (ValueError, IndexError):
            pass
    return None


def detect_majority_year(dates: list) -> "int | None":
    years = [d.year for d in dates if isinstance(d, date)]
    if not years:
        return None
    return Counter(years).most_common(1)[0][0]


def normalize_dates(dates: list) -> list:
    majority = detect_majority_year(dates)
    if majority is None:
        return dates
    fixed = []
    for d in dates:
        if d is None:
            fixed.append(None)
            continue
        if abs(d.year - majority) > 5:
            try:
                fixed.append(d.replace(year=majority))
            except ValueError:
                fixed.append(d)
        else:
            fixed.append(d)
    return fixed


def fmt_date(d) -> str:
    return d.strftime("%d/%m/%Y") if isinstance(d, (date, datetime)) else "—"


def fmt_text(value) -> str:
    return str(value).strip() if value not in (None, "") else "—"


# ─── CONVERSÃO DE ARQUIVO LEGADO ──────────────────────────────
def ensure_xlsx(path: str, should_cancel=None) -> str:
    check_cancel(should_cancel)
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".xlsx":
        return path
    if suffix == ".xls":
        try:
            import xlrd
            from openpyxl import Workbook as _WB
        except ImportError as exc:
            raise ValueError(
                "Arquivos .xls antigos precisam da dependencia xlrd. "
                "Instale as dependencias do requirements.txt ou salve o arquivo como .xlsx."
            ) from exc
        rb = xlrd.open_workbook(path)
        ws_src = rb.sheet_by_index(0)
        wb_dst = _WB()
        try:
            ws_dst = wb_dst.active
            for row_i in range(ws_src.nrows):
                if row_i % 128 == 0:
                    check_cancel(should_cancel)
                for col_i in range(ws_src.ncols):
                    cell = ws_src.cell(row_i, col_i)
                    if cell.ctype == 3:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(cell.value, rb.datemode)
                            ws_dst.cell(row_i + 1, col_i + 1, datetime(*dt_tuple).date())
                        except Exception:
                            ws_dst.cell(row_i + 1, col_i + 1, cell.value)
                    else:
                        ws_dst.cell(row_i + 1, col_i + 1, cell.value)
            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tf.close()
            wb_dst.save(tf.name)
            return tf.name
        finally:
            wb_dst.close()
    # Fallback: tenta abrir com openpyxl (funciona com .xlsm e variantes)
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tf.close()
    try:
        wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
        wb.save(tf.name)
        wb.close()
        check_cancel(should_cancel)
        return tf.name
    except Exception:
        try:
            os.unlink(tf.name)
        except OSError:
            pass
    return path


# ─── DETECÇÃO DE COLUNAS ──────────────────────────────────────
FIELD_ALIASES = {
    "quarto": {
        "quarto", "n_quarto", "nquarto", "numero", "numeroquarto", "room", "apto",
        "apartamento", "apt", "nro", "nr", "suite", "n quarto",
    },
    "nome": {
        "nome", "name", "hospede", "hóspede", "guest", "cliente", "passageiro",
        "ocupante", "nameformatado", "nome_formatado", "nomeformatado",
    },
    "checkin": {
        "check_in", "checkin", "entrada", "check in", "dtcheckin", "dataentrada",
        "data entrada", "dt entrada", "in", "chegada",
    },
    "checkout": {
        "check_out", "checkout", "saida", "saída", "check out", "dtcheckout",
        "datasaida", "data saida", "dt saida", "out", "partida",
    },
}


def _slug(s: str) -> str:
    return strip_accents(_RE_SLUG_SEPARATORS.sub("", str(s).lower().strip()))


_ALIAS_MAP: dict[str, str] = {}
for _field, _aliases in FIELD_ALIASES.items():
    for _a in _aliases:
        _ALIAS_MAP[_slug(_a)] = _field


def detect_columns(ws) -> "dict | None":
    max_col = ws.max_column or 10
    max_row = min(5, ws.max_row or 1)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True), 1
    ):
        found: dict[str, int] = {}
        for ci, cell in enumerate(row):
            slug = _slug(str(cell or ""))
            if slug in _ALIAS_MAP:
                field = _ALIAS_MAP[slug]
                if field not in found:
                    found[field] = ci
        if "nome" in found:
            return {"header_row": row_idx, **found}
    return None


# ─── LEITURA DE ABA ───────────────────────────────────────────
def read_sheet(ws, label: str, should_cancel=None) -> "tuple[dict, list]":
    warnings: list[str] = []
    col_info = detect_columns(ws)
    if col_info is None:
        col_info = {"header_row": 0, "quarto": 0, "nome": 1, "checkin": 2, "checkout": 3}

    header_row = col_info["header_row"]
    ci_quarto  = col_info.get("quarto")
    ci_nome    = col_info.get("nome")
    ci_checkin = col_info.get("checkin")
    ci_checkout= col_info.get("checkout")

    if ci_nome is None:
        raise ValueError(
            f"[{label}] Coluna 'Nome' não encontrada.\n"
            f"Certifique-se de que a planilha tem colunas:\n"
            f"Quarto | Nome | Check-in | Check-out"
        )
    missing_labels = {
        "quarto": "Quarto",
        "checkin": "Check-in",
        "checkout": "Check-out",
    }
    for field, field_label in missing_labels.items():
        if col_info.get(field) is None:
            warnings.append(
                f"[{label}] Coluna '{field_label}' não encontrada. "
                "Os valores correspondentes serão tratados como ausentes."
            )

    # Calcula apenas sobre índices que foram detectados, evitando None silencioso
    col_indices = [c for c in (ci_quarto, ci_nome, ci_checkin, ci_checkout) if c is not None]
    needed_cols = max(col_indices) + 1 if col_indices else 4

    def gcell(idx, row):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    raw_rows: list[dict] = []
    for row_idx, row_vals in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=needed_cols, values_only=True),
        start=header_row + 1,
    ):
        if row_idx % 256 == 0:
            check_cancel(should_cancel)
        raw_nome = gcell(ci_nome, row_vals)
        if not raw_nome or str(raw_nome).strip() in ("", "None"):
            continue
        key   = normalize_key(str(raw_nome))
        match_key = normalizar(str(raw_nome))
        nome  = display_name(str(raw_nome))
        quarto_raw = gcell(ci_quarto, row_vals)
        ci_raw     = gcell(ci_checkin, row_vals)
        co_raw     = gcell(ci_checkout, row_vals)
        quarto = str(quarto_raw).strip() if quarto_raw not in (None, "") else None
        raw_rows.append({
            "key": key, "match_key": match_key, "nome": nome, "quarto": quarto,
            "ci_raw": ci_raw, "co_raw": co_raw, "row_idx": row_idx,
        })

    key_counter = Counter(r["match_key"] for r in raw_rows)
    # Duplicatas por contagem + placeholders sempre como REPETIDO (mesmo únicos)
    dup_keys = {r["match_key"] for r in raw_rows if key_counter[r["match_key"]] > 1}
    dup_keys |= {r["match_key"] for r in raw_rows if r["match_key"] in PLACEHOLDER_NAMES}
    if dup_keys:
        dup_names = sorted({r["nome"] for r in raw_rows if r["match_key"] in dup_keys})
        warnings.append(
            f"[{label}] Nomes repetidos/placeholders marcados como REPETIDO: "
            + ", ".join(dup_names)
        )
    # Garante chaves únicas para duplicatas usando sufixo __dup_N
    dup_seen: dict[str, int] = {}
    for r in raw_rows:
        if r["match_key"] in dup_keys:
            count = dup_seen.get(r["key"], 0)
            dup_seen[r["key"]] = count + 1
            if count > 0:
                r["key"] = f"{r['key']}__dup_{count}"
            r["is_dup"] = True
        else:
            r["is_dup"] = False

    ci_dates = normalize_dates([parse_date_raw(r["ci_raw"]) for r in raw_rows])
    co_dates = normalize_dates([parse_date_raw(r["co_raw"]) for r in raw_rows])

    for i, r in enumerate(raw_rows):
        if ci_dates[i] is None and r["ci_raw"] not in (None, ""):
            warnings.append(
                f"[{label}] Linha {r['row_idx']}: check-in inválido "
                f"'{r['ci_raw']}' para '{r['nome']}'. Use DD/MM/AAAA."
            )
        if co_dates[i] is None and r["co_raw"] not in (None, ""):
            warnings.append(
                f"[{label}] Linha {r['row_idx']}: check-out inválido "
                f"'{r['co_raw']}' para '{r['nome']}'. Use DD/MM/AAAA."
            )

    records: dict[str, dict] = {}
    for i, r in enumerate(raw_rows):
        records[r["key"]] = {
            "nome": r["nome"], "key": r["key"], "match_key": r["match_key"], "quarto": r["quarto"],
            "check_in": ci_dates[i], "check_out": co_dates[i],
            "is_dup": r["is_dup"],
        }

    return records, warnings


# ─── COMPARAÇÃO ───────────────────────────────────────────────
def _make_dup_row(r: dict, fonte: str) -> dict:
    """Gera uma linha REPETIDO com dados apenas da planilha de origem."""
    dash = "—"
    is_sys = fonte == "SISTEMA"
    audit = _empty_match_audit(
        nome_sistema=r["nome"] if is_sys else "",
        nome_hotel=r["nome"] if not is_sys else "",
        decision="REPETIDO",
        stage="DUPLICATA_PLACEHOLDER",
        reason="Registro excluído do matching por nome duplicado ou placeholder.",
    )
    return {
        "nome":    r["nome"],
        "fonte":   fonte,
        "no_match": False,
        "is_dup":  True,
        "room_check_ignored": False,
        "q_sys":   fmt_text(r["quarto"]) if is_sys else dash,
        "q_hotel": dash if is_sys else fmt_text(r["quarto"]),
        "ci_sys":  fmt_date(r["check_in"])  if is_sys else "—",
        "ci_hotel":"—" if is_sys else fmt_date(r["check_in"]),
        "co_sys":  fmt_date(r["check_out"]) if is_sys else "—",
        "co_hotel":"—" if is_sys else fmt_date(r["check_out"]),
        "s_quarto": Status.REPETIDO,
        "s_ci":     Status.REPETIDO,
        "s_co":     Status.REPETIDO,
        "s_geral":  Status.REPETIDO,
        "match_audit": audit,
    }


def _is_room_ok(status: str) -> bool:
    return status in (Status.OK, Status.IGNORADO)


def compare(sys_r, hotel_r, ignorar_quarto: bool = False) -> dict:
    not_found = "NÃO ENCONTRADO"
    no_match  = sys_r is None or hotel_r is None
    q_s_raw = sys_r["quarto"]   if sys_r   else None
    q_h_raw = hotel_r["quarto"] if hotel_r else None
    q_s  = fmt_text(q_s_raw)    if sys_r   else not_found
    q_h  = fmt_text(q_h_raw)    if hotel_r else not_found
    ci_s = sys_r["check_in"]    if sys_r   else None
    ci_h = hotel_r["check_in"]  if hotel_r else None
    co_s = sys_r["check_out"]   if sys_r   else None
    co_h = hotel_r["check_out"] if hotel_r else None

    def _st(v_sys, v_hotel, label):
        if no_match:
            return Status.SEM_CORRESPONDENCIA
        if v_sys is None or v_hotel is None:
            return Status.DATA_AUSENTE
        if v_sys == v_hotel:
            return Status.OK
        return label

    s_q = _st(q_s_raw, q_h_raw, Status.DIVERGENTE)
    room_check_ignored = bool(ignorar_quarto and not no_match)
    if room_check_ignored:
        s_q = Status.IGNORADO
    s_i = _st(ci_s, ci_h, Status.CHECK_IN)
    s_o = _st(co_s, co_h, Status.CHECK_OUT)

    if no_match:
        s_geral = Status.SEM_CORRESPONDENCIA
    elif _is_room_ok(s_q) and s_i == Status.OK and s_o == Status.OK:
        s_geral = Status.OK
    elif any(s == Status.DATA_AUSENTE for s in (s_q, s_i, s_o)):
        s_geral = Status.DATA_AUSENTE
    else:
        divs = []
        if not _is_room_ok(s_q):
            divs.append("QUARTO")
        if s_i != Status.OK:
            divs.append("CHECK-IN")
        if s_o != Status.OK:
            divs.append("CHECK-OUT")
        s_geral = " · ".join(divs)

    audit = (hotel_r or sys_r or {}).get("_match_audit")
    if audit is None:
        audit = _empty_match_audit(
            nome_sistema=sys_r["nome"] if sys_r else "",
            nome_hotel=hotel_r["nome"] if hotel_r else "",
            decision="SEM_MATCH" if no_match else "MATCH",
            stage="SEM_MATCH" if no_match else "SEM_LOG",
            reason=(
                "Registro sem rastreabilidade de matching."
                if no_match else "Match sem rastreabilidade detalhada."
            ),
        )

    return {
        "nome":     (sys_r or hotel_r)["nome"],
        "fonte":    "SISTEMA" if sys_r else "HOTEL",
        "no_match": no_match,
        "is_dup":   False,
        "room_check_ignored": room_check_ignored,
        "q_sys":   q_s,  "q_hotel":  q_h,
        "ci_sys":  fmt_date(ci_s), "ci_hotel": fmt_date(ci_h),
        "co_sys":  fmt_date(co_s), "co_hotel": fmt_date(co_h),
        "s_quarto": s_q,
        "s_ci":     s_i,
        "s_co":     s_o,
        "s_geral":  s_geral,
        "match_audit": audit,
    }


def _apply_name_matching(
    sys_rec: dict[str, dict],
    hotel_rec: dict[str, dict],
    threshold: float = NAME_MATCH_THRESHOLD,
    progress=None,
    should_cancel=None,
) -> dict[str, dict]:
    """Alinha registros únicos do hotel aos nomes únicos do sistema usando fuzzy match 1:1."""
    from .matching import (
        atribuir_indices_matches,
        tokens_significativos,
    )

    sys_items = [(key, rec) for key, rec in sys_rec.items() if not rec["is_dup"]]
    hotel_items = [(key, rec) for key, rec in hotel_rec.items() if not rec["is_dup"]]

    reference_names = [rec["nome"] for _, rec in sys_items]
    reference_normalized = [normalizar(name) for name in reference_names]
    reference_tokens = [set(tokens_significativos(name)) for name in reference_normalized]
    hotel_names = [rec["nome"] for _, rec in hotel_items]
    hotel_normalized = [normalizar(name) for name in hotel_names]

    def matching_progress(fraction: float) -> None:
        if progress is not None:
            progress(55 + round(27 * fraction), "Comparando nomes...")

    matching_details = [{} for _ in hotel_items]
    for _, rec in sys_items:
        rec["_match_audit"] = _empty_match_audit(
            nome_sistema=rec["nome"],
            reason="Nenhum registro da Planilha 2 foi associado a este nome.",
        )

    assignment_indices = atribuir_indices_matches(
        hotel_normalized,
        reference_normalized,
        reference_tokens,
        float(threshold),
        progress=matching_progress,
        should_cancel=should_cancel,
        min_cross_identity_score=CROSS_IDENTITY_MIN_SCORE,
        match_details=matching_details,
    )

    aligned = {key: rec for key, rec in hotel_rec.items() if rec["is_dup"]}

    for index, (original_key, rec) in enumerate(hotel_items):
        if index % 256 == 0:
            check_cancel(should_cancel)
        assignment_index = assignment_indices[index]
        aligned_key = sys_items[assignment_index][0] if assignment_index >= 0 else original_key
        rec["key"] = aligned_key
        audit = _build_match_audit(
            rec["nome"],
            sys_items,
            matching_details[index],
            float(threshold),
            hotel_items=hotel_items,
        )
        rec["_match_audit"] = audit
        if assignment_index >= 0:
            sys_items[assignment_index][1]["_match_audit"] = audit
        aligned[aligned_key] = rec

    return aligned


def run_bate(
    sys_rec: dict,
    hotel_rec: dict,
    ignorar_quarto: bool = False,
    should_cancel=None,
) -> list:
    """Combina as duas fontes: dups geram linhas individuais, demais fazem bate 1:1."""
    results = []
    visited = set()

    # 1. Registros do sistema primeiro
    for index, (k, sys_r) in enumerate(sys_rec.items()):
        if index % 256 == 0:
            check_cancel(should_cancel)
        if sys_r["is_dup"]:
            results.append(_make_dup_row(sys_r, "SISTEMA"))
            continue
        visited.add(k)
        hotel_r = hotel_rec.get(k)
        results.append(compare(sys_r, hotel_r, ignorar_quarto=ignorar_quarto))

    # 2. Registros do hotel que não estavam no sistema
    for index, (k, hotel_r) in enumerate(hotel_rec.items()):
        if index % 256 == 0:
            check_cancel(should_cancel)
        if k in visited:
            continue
        if hotel_r["is_dup"]:
            results.append(_make_dup_row(hotel_r, "HOTEL"))
        else:
            results.append(compare(None, hotel_r, ignorar_quarto=ignorar_quarto))

    return results


def calc_kpis(results: list, should_cancel=None) -> dict:
    """Calcula todos os KPIs em uma única passagem sobre a lista."""
    total = len(results)
    nomap = ok = div = q_div = ci_div = co_div = matched = dups = 0
    for index, r in enumerate(results):
        if index % 256 == 0:
            check_cancel(should_cancel)
        # dups e no_match são categorias exclusivas — verificar separadamente
        if r.get("is_dup") or r.get("s_quarto") == Status.REPETIDO:
            dups += 1
            continue
        if r["no_match"]:
            nomap += 1
            continue
        matched += 1
        all_ok = _is_room_ok(r["s_quarto"]) and r["s_ci"] == Status.OK and r["s_co"] == Status.OK
        if all_ok:
            ok += 1
        else:
            div += 1
        if not _is_room_ok(r["s_quarto"]):
            q_div += 1
        if r["s_ci"] != Status.OK:
            ci_div += 1
        if r["s_co"] != Status.OK:
            co_div += 1
    return dict(total=total, nomap=nomap, dups=dups, matched=matched,
                ok=ok, div=div, q_div=q_div, ci_div=ci_div, co_div=co_div)

def processar_arquivos(
    path1: str,
    path2: str,
    ignorar_quarto: bool = False,
    progress=None,
    should_cancel=None,
) -> "tuple[list, list, dict]":
    """
    Lê as duas planilhas, compara e retorna (results, warnings, kpis).
    Levanta ValueError em caso de problema.
    Temp files gerados por ensure_xlsx são sempre deletados ao final.
    """
    p1 = None
    p2 = None
    wb1 = None
    wb2 = None

    def report(percent: int, message: str) -> None:
        if progress is not None:
            progress(percent, message)

    try:
        report(5, "Preparando planilhas...")
        check_cancel(should_cancel)
        p1 = ensure_xlsx(path1, should_cancel=should_cancel)
        report(14, "Preparando a segunda planilha...")
        p2 = ensure_xlsx(path2, should_cancel=should_cancel)
        report(22, "Abrindo planilhas...")
        wb1 = openpyxl.load_workbook(p1, data_only=True, read_only=True)
        wb2 = openpyxl.load_workbook(p2, data_only=True, read_only=True)
        report(32, "Lendo Planilha 1...")
        sys_rec,   w1 = read_sheet(wb1.active, "Planilha 1", should_cancel=should_cancel)
        report(45, "Lendo Planilha 2...")
        hotel_rec, w2 = read_sheet(wb2.active, "Planilha 2", should_cancel=should_cancel)

        # Cross-contamination: nome dup em qualquer planilha contamina ambas
        report(52, "Identificando nomes repetidos...")
        dup_bases = {
            v["match_key"] for v in sys_rec.values() if v["is_dup"]
        } | {
            v["match_key"] for v in hotel_rec.values() if v["is_dup"]
        }
        for rec in (sys_rec, hotel_rec):
            for index, v in enumerate(rec.values()):
                if index % 256 == 0:
                    check_cancel(should_cancel)
                if v["match_key"] in dup_bases:
                    v["is_dup"] = True

        report(55, "Comparando nomes...")
        hotel_rec = _apply_name_matching(
            sys_rec,
            hotel_rec,
            progress=report,
            should_cancel=should_cancel,
        )
        check_cancel(should_cancel)
        report(85, "Consolidando resultado...")
        results  = run_bate(
            sys_rec,
            hotel_rec,
            ignorar_quarto=ignorar_quarto,
            should_cancel=should_cancel,
        )
        kpis     = calc_kpis(results, should_cancel=should_cancel)
        warnings = w1 + w2
        report(92, "Conferência concluída.")
        return results, warnings, kpis
    finally:
        for wb in (wb1, wb2):
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
        for tmp_path in (p1, p2):
            if tmp_path and tmp_path not in (path1, path2):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass



