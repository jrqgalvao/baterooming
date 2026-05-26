"""
bate_rooming_core.py — Bate-Rooming v5 (Your Company)
MÃ³dulo de lÃ³gica pura: leitura, comparaÃ§Ã£o e exportaÃ§Ã£o Excel.
Sem dependÃªncias de UI (sem Tkinter).
"""

import os
import re
import tempfile
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_RE_SPACES = re.compile(r"\s+")
_RE_SLUG_SEPARATORS = re.compile(r"[\s_\-\./]+")

# â”€â”€â”€ STATUS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# StrEnum foi adicionado no Python 3.11; fallback limpo para 3.10.
try:
    from enum import StrEnum
except ImportError:
    from enum import Enum

    class StrEnum(str, Enum):  # type: ignore
        pass


class Status(StrEnum):
    OK = "OK"
    DIVERGENTE = "DIVERGENTE"
    IGNORADO = "IGNORADO"
    CHECK_IN = "CHECK-IN"
    CHECK_OUT = "CHECK-OUT"
    DATA_AUSENTE = "DATA AUSENTE"
    SEM_CORRESPONDENCIA = "SEM CORRESPONDÊNCIA"
    A_VERIFICAR = "A VERIFICAR"
    REPETIDO = "REPETIDO"


# â”€â”€â”€ UTILITÃRIOS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", str(text))
        if unicodedata.category(c) != "Mn"
    )


def normalize_key(text: str) -> str:
    clean = _RE_SPACES.sub(" ", str(text or "").strip())
    return strip_accents(clean).lower()



# Nomes placeholder que devem sempre ser tratados como REPETIDO
PLACEHOLDER_NAMES = {
    "a nomear", "a definir", "tbd", "to be defined",
    "to be confirmed", "tbc", "sem nome", "n/a", "none",
    "anomear", "adefinir", "semhospede", "sem hospede",
}

def display_name(text: str) -> str:
    return " ".join(
        w.capitalize()
        for w in _RE_SPACES.sub(" ", str(text or "").strip()).split()
    )


_FMT_LENGTHS = {
    "%d/%m/%Y": 10, "%d/%m/%y": 8,
    "%Y-%m-%d %H:%M:%S": 19, "%Y-%m-%d": 10,
    "%d-%m-%Y": 10, "%d.%m.%Y": 10, "%m/%d/%Y": 10,
}


def parse_date_raw(value) -> "date | None":
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s or s in ("None", "—", "-", ""):
        return None
    for fmt, length in _FMT_LENGTHS.items():
        try:
            return datetime.strptime(s[:length], fmt).date()
        except (ValueError, IndexError):
            pass
    return None


def detect_majority_year(dates: list) -> "int | None":
    years = [d.year for d in dates if isinstance(d, date)]
    if not years:
        return None
    return Counter(years).most_common(1)[0][0]


def normalize_dates(dates: list) -> list:
    majority = detect_majority_year(dates)
    if majority is None:
        return dates
    fixed = []
    for d in dates:
        if d is None:
            fixed.append(None)
            continue
        if abs(d.year - majority) > 5:
            try:
                fixed.append(d.replace(year=majority))
            except ValueError:
                fixed.append(d)
        else:
            fixed.append(d)
    return fixed


def fmt_date(d) -> str:
    return d.strftime("%d/%m/%Y") if isinstance(d, (date, datetime)) else "—"


def fmt_text(value) -> str:
    return str(value).strip() if value not in (None, "") else "—"


# â”€â”€â”€ CONVERSÃƒO DE ARQUIVO LEGADO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def ensure_xlsx(path: str) -> str:
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".xlsx":
        return path
    if suffix == ".xls":
        try:
            import xlrd
            from openpyxl import Workbook as _WB
        except ImportError as exc:
            raise ValueError(
                "Arquivos .xls antigos precisam da dependencia xlrd. "
                "Instale as dependencias do requirements.txt ou salve o arquivo como .xlsx."
            ) from exc
        try:
            rb = xlrd.open_workbook(path)
            ws_src = rb.sheet_by_index(0)
            wb_dst = _WB()
            ws_dst = wb_dst.active
            for row_i in range(ws_src.nrows):
                for col_i in range(ws_src.ncols):
                    cell = ws_src.cell(row_i, col_i)
                    if cell.ctype == 3:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(cell.value, rb.datemode)
                            ws_dst.cell(row_i + 1, col_i + 1, datetime(*dt_tuple).date())
                        except Exception:
                            ws_dst.cell(row_i + 1, col_i + 1, cell.value)
                    else:
                        ws_dst.cell(row_i + 1, col_i + 1, cell.value)
            tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tf.close()
            wb_dst.save(tf.name)
            return tf.name
        except ValueError:
            raise
    # Fallback: tenta abrir com openpyxl (funciona com .xlsm e variantes)
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tf.close()
    try:
        wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
        wb.save(tf.name)
        wb.close()
        return tf.name
    except Exception:
        try:
            os.unlink(tf.name)
        except OSError:
            pass
    return path


# â”€â”€â”€ DETECÃ‡ÃƒO DE COLUNAS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
FIELD_ALIASES = {
    "quarto": {
        "quarto", "n_quarto", "nquarto", "numero", "numeroquarto", "room", "apto",
        "apartamento", "apt", "nro", "nr", "suite", "n quarto",
    },
    "nome": {
        "nome", "name", "hospede", "hóspede", "guest", "cliente", "passageiro",
        "ocupante", "nameformatado", "nome_formatado", "nomeformatado",
    },
    "checkin": {
        "check_in", "checkin", "entrada", "check in", "dtcheckin", "dataentrada",
        "data entrada", "dt entrada", "in", "chegada",
    },
    "checkout": {
        "check_out", "checkout", "saida", "saída", "check out", "dtcheckout",
        "datasaida", "data saida", "dt saida", "out", "partida",
    },
}


def _slug(s: str) -> str:
    return strip_accents(_RE_SLUG_SEPARATORS.sub("", str(s).lower().strip()))


_ALIAS_MAP: dict[str, str] = {}
for _field, _aliases in FIELD_ALIASES.items():
    for _a in _aliases:
        _ALIAS_MAP[_slug(_a)] = _field


def detect_columns(ws) -> "dict | None":
    max_col = ws.max_column or 10
    max_row = min(5, ws.max_row or 1)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True), 1
    ):
        found: dict[str, int] = {}
        for ci, cell in enumerate(row):
            slug = _slug(str(cell or ""))
            if slug in _ALIAS_MAP:
                field = _ALIAS_MAP[slug]
                if field not in found:
                    found[field] = ci
        if "nome" in found:
            return {"header_row": row_idx, **found}
    return None


# â”€â”€â”€ LEITURA DE ABA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def read_sheet(ws, label: str) -> "tuple[dict, list]":
    warnings: list[str] = []
    col_info = detect_columns(ws)
    if col_info is None:
        col_info = {"header_row": 0, "quarto": 0, "nome": 1, "checkin": 2, "checkout": 3}

    header_row = col_info["header_row"]
    ci_quarto  = col_info.get("quarto")
    ci_nome    = col_info.get("nome")
    ci_checkin = col_info.get("checkin")
    ci_checkout= col_info.get("checkout")

    if ci_nome is None:
        raise ValueError(
            f"[{label}] Coluna 'Nome' não encontrada.\n"
            f"Certifique-se de que a planilha tem colunas:\n"
            f"Quarto | Nome | Check-in | Check-out"
        )
    missing_labels = {
        "quarto": "Quarto",
        "checkin": "Check-in",
        "checkout": "Check-out",
    }
    for field, field_label in missing_labels.items():
        if col_info.get(field) is None:
            warnings.append(
                f"[{label}] Coluna '{field_label}' não encontrada. "
                "Os valores correspondentes serão tratados como ausentes."
            )

    # Calcula apenas sobre Ã­ndices que foram detectados, evitando None silencioso
    col_indices = [c for c in (ci_quarto, ci_nome, ci_checkin, ci_checkout) if c is not None]
    needed_cols = max(col_indices) + 1 if col_indices else 4

    def gcell(idx, row):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    raw_rows: list[dict] = []
    for row_idx, row_vals in enumerate(
        ws.iter_rows(min_row=header_row + 1, max_col=needed_cols, values_only=True),
        start=header_row + 1,
    ):
        raw_nome = gcell(ci_nome, row_vals)
        if not raw_nome or str(raw_nome).strip() in ("", "None"):
            continue
        key   = normalize_key(str(raw_nome))
        nome  = display_name(str(raw_nome))
        quarto_raw = gcell(ci_quarto, row_vals)
        ci_raw     = gcell(ci_checkin, row_vals)
        co_raw     = gcell(ci_checkout, row_vals)
        quarto = str(quarto_raw).strip() if quarto_raw not in (None, "") else None
        raw_rows.append({
            "key": key, "nome": nome, "quarto": quarto,
            "ci_raw": ci_raw, "co_raw": co_raw, "row_idx": row_idx,
        })

    key_counter = Counter(r["key"] for r in raw_rows)
    # Duplicatas por contagem + placeholders sempre como REPETIDO (mesmo Ãºnicos)
    dup_keys = {r["key"] for r in raw_rows if key_counter[r["key"]] > 1}
    dup_keys |= {r["key"] for r in raw_rows if r["key"] in PLACEHOLDER_NAMES}
    if dup_keys:
        dup_names = sorted({r["nome"] for r in raw_rows if r["key"] in dup_keys})
        warnings.append(
            f"[{label}] Nomes repetidos/placeholders marcados como REPETIDO: "
            + ", ".join(dup_names)
        )
    # Garante chaves Ãºnicas para duplicatas usando sufixo __dup_N
    dup_seen: dict[str, int] = {}
    for r in raw_rows:
        if r["key"] in dup_keys:
            count = dup_seen.get(r["key"], 0)
            dup_seen[r["key"]] = count + 1
            if count > 0:
                r["key"] = f"{r['key']}__dup_{count}"
            r["is_dup"] = True
        else:
            r["is_dup"] = False

    ci_dates = normalize_dates([parse_date_raw(r["ci_raw"]) for r in raw_rows])
    co_dates = normalize_dates([parse_date_raw(r["co_raw"]) for r in raw_rows])

    for i, r in enumerate(raw_rows):
        if ci_dates[i] is None and r["ci_raw"] not in (None, ""):
            warnings.append(
                f"[{label}] Linha {r['row_idx']}: check-in inválido "
                f"'{r['ci_raw']}' para '{r['nome']}'. Use DD/MM/AAAA."
            )
        if co_dates[i] is None and r["co_raw"] not in (None, ""):
            warnings.append(
                f"[{label}] Linha {r['row_idx']}: check-out inválido "
                f"'{r['co_raw']}' para '{r['nome']}'. Use DD/MM/AAAA."
            )

    records: dict[str, dict] = {}
    for i, r in enumerate(raw_rows):
        records[r["key"]] = {
            "nome": r["nome"], "key": r["key"], "quarto": r["quarto"],
            "check_in": ci_dates[i], "check_out": co_dates[i],
            "is_dup": r["is_dup"],
        }

    return records, warnings


# â”€â”€â”€ COMPARAÃ‡ÃƒO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _make_dup_row(r: dict, fonte: str) -> dict:
    """Gera uma linha REPETIDO com dados apenas da planilha de origem."""
    dash = "—"
    is_sys = fonte == "SISTEMA"
    return {
        "nome":    r["nome"],
        "fonte":   fonte,
        "no_match": False,
        "is_dup":  True,
        "room_check_ignored": False,
        "q_sys":   fmt_text(r["quarto"]) if is_sys else dash,
        "q_hotel": dash if is_sys else fmt_text(r["quarto"]),
        "ci_sys":  fmt_date(r["check_in"])  if is_sys else "—",
        "ci_hotel":"—" if is_sys else fmt_date(r["check_in"]),
        "co_sys":  fmt_date(r["check_out"]) if is_sys else "—",
        "co_hotel":"—" if is_sys else fmt_date(r["check_out"]),
        "s_quarto": Status.REPETIDO,
        "s_ci":     Status.REPETIDO,
        "s_co":     Status.REPETIDO,
        "s_geral":  Status.REPETIDO,
    }


def _is_room_ok(status: str) -> bool:
    return status in (Status.OK, Status.IGNORADO)


def compare(sys_r, hotel_r, ignorar_quarto: bool = False) -> dict:
    not_found = "NÃO ENCONTRADO"
    no_match  = sys_r is None or hotel_r is None
    q_s_raw = sys_r["quarto"]   if sys_r   else None
    q_h_raw = hotel_r["quarto"] if hotel_r else None
    q_s  = fmt_text(q_s_raw)    if sys_r   else not_found
    q_h  = fmt_text(q_h_raw)    if hotel_r else not_found
    ci_s = sys_r["check_in"]    if sys_r   else None
    ci_h = hotel_r["check_in"]  if hotel_r else None
    co_s = sys_r["check_out"]   if sys_r   else None
    co_h = hotel_r["check_out"] if hotel_r else None

    def _st(v_sys, v_hotel, label):
        if no_match:
            return Status.SEM_CORRESPONDENCIA
        if v_sys is None or v_hotel is None:
            return Status.DATA_AUSENTE
        if v_sys == v_hotel:
            return Status.OK
        return label

    s_q = _st(q_s_raw, q_h_raw, Status.DIVERGENTE)
    room_check_ignored = bool(ignorar_quarto and not no_match)
    if room_check_ignored:
        s_q = Status.IGNORADO
    s_i = _st(ci_s, ci_h, Status.CHECK_IN)
    s_o = _st(co_s, co_h, Status.CHECK_OUT)

    if no_match:
        s_geral = Status.SEM_CORRESPONDENCIA
    elif _is_room_ok(s_q) and s_i == Status.OK and s_o == Status.OK:
        s_geral = Status.OK
    elif any(s == Status.DATA_AUSENTE for s in (s_q, s_i, s_o)):
        s_geral = Status.DATA_AUSENTE
    else:
        divs = []
        if not _is_room_ok(s_q): divs.append("QUARTO")
        if s_i != Status.OK: divs.append("CHECK-IN")
        if s_o != Status.OK: divs.append("CHECK-OUT")
        s_geral = " · ".join(divs)

    return {
        "nome":     (sys_r or hotel_r)["nome"],
        "fonte":    "SISTEMA" if sys_r else "HOTEL",
        "no_match": no_match,
        "is_dup":   False,
        "room_check_ignored": room_check_ignored,
        "q_sys":   q_s,  "q_hotel":  q_h,
        "ci_sys":  fmt_date(ci_s), "ci_hotel": fmt_date(ci_h),
        "co_sys":  fmt_date(co_s), "co_hotel": fmt_date(co_h),
        "s_quarto": s_q,
        "s_ci":     s_i,
        "s_co":     s_o,
        "s_geral":  s_geral,
    }


def run_bate(sys_rec: dict, hotel_rec: dict, ignorar_quarto: bool = False) -> list:
    """Combina as duas fontes: dups geram linhas individuais, demais fazem bate 1:1."""
    results = []
    visited = set()

    # 1. Registros do sistema primeiro
    for k, sys_r in sys_rec.items():
        if sys_r["is_dup"]:
            results.append(_make_dup_row(sys_r, "SISTEMA"))
            continue
        visited.add(k)
        hotel_r = hotel_rec.get(k)
        results.append(compare(sys_r, hotel_r, ignorar_quarto=ignorar_quarto))

    # 2. Registros do hotel que nÃ£o estavam no sistema
    for k, hotel_r in hotel_rec.items():
        if k in visited:
            continue
        if hotel_r["is_dup"]:
            results.append(_make_dup_row(hotel_r, "HOTEL"))
        else:
            results.append(compare(None, hotel_r, ignorar_quarto=ignorar_quarto))

    return results


def calc_kpis(results: list) -> dict:
    """Calcula todos os KPIs em uma Ãºnica passagem sobre a lista."""
    total = len(results)
    nomap = ok = div = q_div = ci_div = co_div = matched = dups = 0
    for r in results:
        # dups e no_match sÃ£o categorias exclusivas — verificar separadamente
        if r.get("is_dup") or r.get("s_quarto") == Status.REPETIDO:
            dups += 1
            continue
        if r["no_match"]:
            nomap += 1
            continue
        matched += 1
        all_ok = _is_room_ok(r["s_quarto"]) and r["s_ci"] == Status.OK and r["s_co"] == Status.OK
        if all_ok:
            ok += 1
        else:
            div += 1
        if not _is_room_ok(r["s_quarto"]): q_div  += 1
        if r["s_ci"]     != Status.OK: ci_div += 1
        if r["s_co"]     != Status.OK: co_div += 1
    return dict(total=total, nomap=nomap, dups=dups, matched=matched,
                ok=ok, div=div, q_div=q_div, ci_div=ci_div, co_div=co_div)


# â”€â”€â”€ EXPORTAÃ‡ÃƒO EXCEL â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _font(*, bold=False, italic=False, color="162016", size=10) -> Font:
    return Font(bold=bold, italic=italic, color=color.lstrip("#"), size=size)


_THIN = Border(
    left=Side(style="thin", color="C8D8C8"),
    right=Side(style="thin", color="C8D8C8"),
    top=Side(style="thin", color="C8D8C8"),
    bottom=Side(style="thin", color="C8D8C8"),
)

ALIGN_LEFT        = Alignment(vertical="center", horizontal="left")
ALIGN_CENTER      = Alignment(vertical="center", horizontal="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Fills de linha por categoria (para filtrar por cor no Excel)
FILL_ROW_OK     = _fill("F7FBF7")   # verde
FILL_ROW_DIV    = _fill("FFFDF4")   # amarelo
FILL_ROW_NOMAP  = _fill("F8F6FC")   # roxo
FILL_ROW_REP    = _fill("FDF7FA")   # rosa
FILL_ROW_ABS    = _fill("FFFBF5")   # laranja

def _row_fill(r: dict) -> "PatternFill":
    """Retorna o fill correto para a linha inteira baseado na categoria dominante."""
    if r.get("is_dup") or r.get("s_quarto") == "REPETIDO":
        return FILL_ROW_REP
    if r.get("no_match"):
        return FILL_ROW_NOMAP
    statuses = [r.get("s_quarto"), r.get("s_ci"), r.get("s_co")]
    if any(s == Status.DATA_AUSENTE for s in statuses):
        return FILL_ROW_ABS
    if _is_room_ok(r.get("s_quarto")) and r.get("s_ci") == Status.OK and r.get("s_co") == Status.OK:
        return FILL_ROW_OK
    return FILL_ROW_DIV

STYLE_CACHE = {
    "ok":         (_fill("E6F4E7"), _font(color="1E6B22")),
    "nomap":      (_fill("EDE7F6"), _font(bold=True, color="4527A0")),
    "absent":     (_fill("FFF3E0"), _font(bold=True, color="BF360C")),
    "div":        (_fill("FFEBEE"), _font(bold=True, color="B71C1C")),
    "averificar": (_fill("FFF8E1"), _font(bold=True, color="F57F17")),
    "repetido":   (_fill("FCE4EC"), _font(bold=True, color="880E4F")),
    "text":       _font(color="162016"),
    "title":      _font(bold=True, color="FFFFFF", size=11),
    "sum_title":  _font(bold=True, color="FFFFFF", size=13),
    "sum_sub":    _font(italic=True, color="6B8F6B", size=9),
}

_XL_HEADERS = [
    "NOME", "FONTE",
    "QUARTO (PLAN. 1)", "QUARTO (PLAN. 2)",
    "CHECK-IN (PLAN. 1)", "CHECK-IN (PLAN. 2)",
    "CHECK-OUT (PLAN. 1)", "CHECK-OUT (PLAN. 2)",
    "STATUS GERAL", "STATUS QUARTO", "STATUS CHECK-IN", "STATUS CHECK-OUT",
]
_XL_WIDTHS = [38, 10, 18, 18, 20, 20, 20, 20, 28, 22, 22, 22]
_XL_ROOM_COLUMNS = {2, 3, 9}


def _hide_room_columns(results: list) -> bool:
    return any(r.get("room_check_ignored") for r in results)


def _visible_excel_columns(hide_room: bool = False) -> list[int]:
    if not hide_room:
        return list(range(len(_XL_HEADERS)))
    return [idx for idx in range(len(_XL_HEADERS)) if idx not in _XL_ROOM_COLUMNS]


def _status_style(v: str):
    if v == Status.OK:                  return STYLE_CACHE["ok"]
    if v == Status.IGNORADO:            return STYLE_CACHE["averificar"]
    if v == Status.SEM_CORRESPONDENCIA: return STYLE_CACHE["nomap"]
    if v == Status.DATA_AUSENTE:        return STYLE_CACHE["absent"]
    if v == Status.A_VERIFICAR:         return STYLE_CACHE["averificar"]
    if v == Status.REPETIDO:            return STYLE_CACHE["repetido"]
    if "·" in v:                         return STYLE_CACHE["div"]   # status geral composto
    return STYLE_CACHE["div"]


def _set_header_cell(ws, row, col, text, bg="1B5E20", fg="FFFFFF", sz=10):
    c = ws.cell(row, col, text)
    c.font      = Font(bold=True, color=fg, size=sz)
    c.fill      = _fill(bg)
    c.alignment = ALIGN_CENTER_WRAP
    c.border    = _THIN
    return c


def _all_result_vals(r: dict) -> list:
    """Extrai a lista de valores de uma linha de resultado para uso no Excel."""
    return [
        r["nome"], r["fonte"],
        r["q_sys"], r["q_hotel"],
        r["ci_sys"], r["ci_hotel"],
        r["co_sys"], r["co_hotel"],
        r.get("s_geral", r["s_quarto"]), r["s_quarto"], r["s_ci"], r["s_co"],
    ]


def _result_vals(r: dict, hide_room: bool = False) -> list:
    vals = _all_result_vals(r)
    return [vals[idx] for idx in _visible_excel_columns(hide_room)]


def _apply_data_row(ws, ri, vals, row_fill=None, status_start_col: int = 9):
    if row_fill is None:
        row_fill = _fill("FFFFFF")
    base_font = STYLE_CACHE["text"]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v)
        c.border    = _THIN
        c.alignment = ALIGN_LEFT if ci == 1 else ALIGN_CENTER
        if ci >= status_start_col:
            fill, font = _status_style(str(v))
            c.fill = fill
            c.font = font
        else:
            c.fill = row_fill
            c.font = base_font
    ws.row_dimensions[ri].height = 18


def _apply_filter(ws, header_row: int, last_row: int) -> None:
    if last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{last_row}"


def _build_result_sheet(ws, results: list, title: str = "RESULTADO", hide_room: bool = False):
    columns = _visible_excel_columns(hide_room)
    status_start_col = next((i for i, idx in enumerate(columns, 1) if idx >= 8), len(columns) + 1)
    ws.title = title
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    t = ws.cell(1, 1, f"Bate-Rooming Your Company — {title} • Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    t.font      = STYLE_CACHE["title"]
    t.fill      = _fill("141E14")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    for ci, idx in enumerate(columns, 1):
        _set_header_cell(ws, 2, ci, _XL_HEADERS[idx])
        ws.column_dimensions[get_column_letter(ci)].width = _XL_WIDTHS[idx]
    ws.row_dimensions[2].height = 30
    for ri, r in enumerate(results, 3):
        _apply_data_row(
            ws,
            ri,
            _result_vals(r, hide_room=hide_room),
            row_fill=_row_fill(r),
            status_start_col=status_start_col,
        )
    _apply_filter(ws, 2, len(results) + 2)
    # Linha de totalizador por categoria
    kpi = calc_kpis(results)
    footer_row = len(results) + 4
    ws.merge_cells(f"A{footer_row}:{get_column_letter(len(columns))}{footer_row}")
    summary_text = (
        f"TOTAL: {kpi['total']}  |  "
        f"✓ OK: {kpi['ok']}  |  "
        f"⚠ Divergentes: {kpi['div']}  |  "
        f"? Sem correspondência: {kpi['nomap']}  |  "
        f"🔁 Repetidos: {kpi['dups']}"
    )
    fc = ws.cell(footer_row, 1, summary_text)
    fc.font      = Font(bold=True, color="FFFFFF", size=9)
    fc.fill      = _fill("1B3A1B")
    fc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[footer_row].height = 20


def _build_summary_sheet(ws, results: list, hide_room: bool = False):
    ws.title = "RESUMO"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.merge_cells("A1:C1")
    t = ws.cell(1, 1, "Bate-Rooming Your Company — Resumo da Conferência")
    t.font      = STYLE_CACHE["sum_title"]
    t.fill      = _fill("141E14")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:C2")
    d = ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
    d.font      = STYLE_CACHE["sum_sub"]
    d.fill      = _fill("141E14")
    d.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18
    kpi = calc_kpis(results)
    kpis_geral = [
        ("Total de hóspedes",           kpi["total"],   "162016", "F4F6F4", "B8CCB8"),
        ("Verificar manualmente (?)",   kpi["nomap"],   "4527A0", "EDE7F6", "7E57C2"),
        ("Nomes repetidos (REPETIDO)",     kpi["dups"], "880E4F", "FCE4EC", "CE93D8"),
        ("Com correspondência",         kpi["matched"], "162016", "F4F6F4", "B8CCB8"),
        ("Sem divergência ✓",           kpi["ok"],      "1E6B22", "E6F4E7", "1E6B22"),
        ("Com divergência ⚠",           kpi["div"],     "BF360C", "FFF3E0", "BF360C"),
    ]
    kpis_div = []
    if not hide_room:
        kpis_div.append(("Divergência — Quarto", kpi["q_div"], "B71C1C", "FFEBEE", "B71C1C"))
    kpis_div.extend([
        ("Divergência — Check-in",  kpi["ci_div"], "B71C1C", "FFEBEE", "B71C1C"),
        ("Divergência — Check-out", kpi["co_div"], "B71C1C", "FFEBEE", "B71C1C"),
    ])

    def write_section(start_row, section_label, section_color, rows):
        ws.merge_cells(f"A{start_row}:C{start_row}")
        h = ws.cell(start_row, 1, section_label)
        h.font      = Font(bold=True, color="FFFFFF", size=10)
        h.fill      = _fill(section_color)
        h.alignment = ALIGN_CENTER
        ws.row_dimensions[start_row].height = 22
        _set_header_cell(ws, start_row + 1, 1, "MÉTRICA",    section_color)
        _set_header_cell(ws, start_row + 1, 2, "QUANTIDADE", section_color)
        _set_header_cell(ws, start_row + 1, 3, "% DO TOTAL", section_color)
        ws.row_dimensions[start_row + 1].height = 20
        for i, (label, val, fg, bg, brd) in enumerate(rows, start_row + 2):
            row_fill = _fill(bg)
            for ci in range(1, 4):
                cell = ws.cell(i, ci)
                cell.fill = row_fill
                cell.border = _THIN
            lbl_c           = ws.cell(i, 1, label)
            lbl_c.font      = Font(color=fg, size=10)
            lbl_c.alignment = ALIGN_LEFT
            val_c           = ws.cell(i, 2, val)
            val_c.font      = Font(bold=True, color=fg, size=12)
            val_c.alignment = ALIGN_CENTER
            pct             = f"{val/kpi['total']*100:.1f}%" if kpi["total"] else "—"
            pct_c           = ws.cell(i, 3, pct)
            pct_c.font      = Font(color=fg, size=10)
            pct_c.alignment = ALIGN_CENTER
            ws.row_dimensions[i].height = 22
        return start_row + 2 + len(rows)

    next_row = write_section(4, "VISÃO GERAL", "162016", kpis_geral)
    next_row += 1
    next_row = write_section(next_row, "DETALHES DE DIVERGÊNCIA", "B71C1C", kpis_div)
    foot_row = next_row + 1
    ws.merge_cells(f"A{foot_row}:C{foot_row}")
    f = ws.cell(foot_row, 1, "Your Company © 2026 — Bate-Rooming  |  Abas: RESULTADO COMPLETO · DIVERGÊNCIAS · SEM CORRESPONDÊNCIA · REPETIDOS")
    f.font      = Font(italic=True, color="9AAA9A", size=8)
    f.alignment = ALIGN_CENTER


def _build_sheet_header(ws, title: str, bg: str, hide_room: bool = False):
    """Cria a linha de tÃ­tulo e configura a aba."""
    columns = _visible_excel_columns(hide_room)
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    t = ws.cell(1, 1, f"{title} • Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    t.font      = STYLE_CACHE["title"]
    t.fill      = _fill(bg)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _build_data_section(ws, rows: list, bg: str, start_row: int, empty_msg: str, hide_room: bool = False) -> int:
    """Adiciona cabeÃ§alho + linhas de dados. Retorna prÃ³xima linha disponÃ­vel."""
    columns = _visible_excel_columns(hide_room)
    status_start_col = next((i for i, idx in enumerate(columns, 1) if idx >= 8), len(columns) + 1)
    cur_row = start_row
    if rows:
        header_row = cur_row
        for ci, idx in enumerate(columns, 1):
            _set_header_cell(ws, cur_row, ci, _XL_HEADERS[idx], bg=bg)
            ws.column_dimensions[get_column_letter(ci)].width = _XL_WIDTHS[idx]
        ws.row_dimensions[cur_row].height = 30
        cur_row += 1
        for r in rows:
            _apply_data_row(
                ws,
                cur_row,
                _result_vals(r, hide_room=hide_room),
                row_fill=_row_fill(r),
                status_start_col=status_start_col,
            )
            cur_row += 1
        _apply_filter(ws, header_row, cur_row - 1)
    else:
        ws.merge_cells(f"A{cur_row}:{get_column_letter(len(columns))}{cur_row}")
        c           = ws.cell(cur_row, 1, empty_msg)
        c.font      = Font(bold=True, color="1E6B22", size=11)
        c.fill      = _fill("E6F4E7")
        c.alignment = ALIGN_CENTER
        ws.row_dimensions[cur_row].height = 30
        cur_row += 1
    return cur_row


def _categorize_export_rows(results: list) -> tuple[list, list, list]:
    divergent = []
    nomap = []
    repeated = []
    for r in results:
        is_repeated = r.get("is_dup") or r.get("s_quarto") == "REPETIDO"
        if is_repeated:
            repeated.append(r)
        elif r["no_match"]:
            nomap.append(r)
        elif not (_is_room_ok(r["s_quarto"]) and r["s_ci"] == Status.OK and r["s_co"] == Status.OK):
            divergent.append(r)
    return divergent, nomap, repeated


def _build_divergences_sheet(ws, divergent: list, hide_room: bool = False):
    ws.title = "DIVERGÊNCIAS"
    _build_sheet_header(ws,
        f"DIVERGÊNCIAS — {len(divergent)} registro(s) com dados divergentes entre as planilhas",
        "B71C1C", hide_room=hide_room)
    _build_data_section(ws, divergent, "B71C1C", 2, "✓ Nenhuma divergência encontrada!", hide_room=hide_room)
    ws.freeze_panes = "A3"


def _build_nomap_sheet(ws, nomap: list, hide_room: bool = False):
    ws.title = "SEM CORRESPONDÊNCIA"
    _build_sheet_header(ws,
        f"SEM CORRESPONDÊNCIA — {len(nomap)} registro(s) sem par entre as planilhas",
        "4527A0", hide_room=hide_room)
    _build_data_section(ws, nomap, "4527A0", 2, "✓ Nenhum registro sem correspondência!", hide_room=hide_room)
    ws.freeze_panes = "A3"


def _build_repeated_sheet(ws, repeated: list, hide_room: bool = False):
    ws.title = "REPETIDOS"
    _build_sheet_header(ws,
        f"REPETIDOS — {len(repeated)} registro(s) com nome duplicado ou placeholder",
        "880E4F", hide_room=hide_room)
    _build_data_section(ws, repeated, "880E4F", 2, "✓ Nenhum nome repetido encontrado!", hide_room=hide_room)
    ws.freeze_panes = "A3"



def write_excel(results: list, path: Path):
    hide_room = _hide_room_columns(results)
    divergent, nomap, repeated = _categorize_export_rows(results)
    wb         = openpyxl.Workbook()
    ws_summary = wb.active
    _build_summary_sheet(ws_summary, results, hide_room=hide_room)
    ws_result  = wb.create_sheet()
    _build_result_sheet(ws_result, results, "RESULTADO COMPLETO", hide_room=hide_room)
    ws_divs    = wb.create_sheet()
    _build_divergences_sheet(ws_divs, divergent, hide_room=hide_room)
    ws_nomap   = wb.create_sheet()
    _build_nomap_sheet(ws_nomap, nomap, hide_room=hide_room)
    ws_rep     = wb.create_sheet()
    _build_repeated_sheet(ws_rep, repeated, hide_room=hide_room)
    wb.active  = ws_summary
    wb.save(path)


# â”€â”€â”€ FUNÃ‡ÃƒO DE ALTO NÃVEL: processar dois arquivos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def processar_arquivos(path1: str, path2: str, ignorar_quarto: bool = False) -> "tuple[list, list, dict]":
    """
    LÃª as duas planilhas, compara e retorna (results, warnings, kpis).
    Levanta ValueError em caso de problema.
    Temp files gerados por ensure_xlsx sÃ£o sempre deletados ao final.
    """
    p1 = ensure_xlsx(path1)
    p2 = ensure_xlsx(path2)
    wb1 = None
    wb2 = None
    try:
        wb1 = openpyxl.load_workbook(p1, data_only=True)
        wb2 = openpyxl.load_workbook(p2, data_only=True)
        sys_rec,   w1 = read_sheet(wb1.active, "Planilha 1")
        hotel_rec, w2 = read_sheet(wb2.active, "Planilha 2")

        # Cross-contamination: nome dup em qualquer planilha contamina ambas
        def _base_key(k: str) -> str:
            return k.split("__dup_")[0]

        dup_bases = {
            _base_key(k) for k, v in sys_rec.items()   if v["is_dup"]
        } | {
            _base_key(k) for k, v in hotel_rec.items() if v["is_dup"]
        }
        for rec in (sys_rec, hotel_rec):
            for k, v in rec.items():
                if _base_key(k) in dup_bases:
                    v["is_dup"] = True

        results  = run_bate(sys_rec, hotel_rec, ignorar_quarto=ignorar_quarto)
        kpis     = calc_kpis(results)
        warnings = w1 + w2
        return results, warnings, kpis
    finally:
        for wb in (wb1, wb2):
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
        for tmp_path in (p1, p2):
            if tmp_path not in (path1, path2):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass


