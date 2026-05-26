import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from match_nomes_core import executar_match, write_excel


def _write_names(path: Path, names: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    for name in names:
        ws.append([name])
    wb.save(path)


def _write_rows(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


class MatchNomesRegressionTests(unittest.TestCase):
    def test_no_header_rooming_layout_reads_names_from_column_b(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "planilha_1.xlsx"
            planilha_2 = base / "planilha_2.xlsx"

            _write_rows(
                planilha_1,
                [
                    ["101", "Ana Maria Oliveira", "01/01/2026", "02/01/2026"],
                    ["102", "Bruno Costa", "01/01/2026", "02/01/2026"],
                ],
            )
            _write_rows(
                planilha_2,
                [
                    ["201", "BRUNO C.", "03/01/2026", "04/01/2026"],
                    ["202", "ANA OLIVEIRA", "05/01/2026", "06/01/2026"],
                ],
            )

            result = executar_match(str(planilha_1), str(planilha_2), 65)

        self.assertTrue(result["ok"], result.get("erro"))
        self.assertEqual(result["nomes_finais"], ["Bruno Costa", "Ana Maria Oliveira"])
        self.assertEqual(result["name_rows"], [1, 2])
        self.assertEqual(result["template_path"], str(planilha_2))

    def test_export_updates_only_planilha_2_name_cells_and_preserves_template(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "planilha_1.xlsx"
            planilha_2 = base / "planilha_2.xlsx"
            output = base / "output.xlsx"

            _write_rows(
                planilha_1,
                [
                    ["room number", "name", "check in date", "check out date"],
                    ["101", "Ana Maria Oliveira", "01/01/2026", "02/01/2026"],
                    ["102", "Bruno Costa", "01/01/2026", "02/01/2026"],
                ],
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(["room number", "name", "check in date", "check out date"])
            ws.append(["201", "BRUNO C.", "03/01/2026", "04/01/2026"])
            ws.append(["202", "ANA OLIVEIRA", "05/01/2026", "06/01/2026"])
            ws.append(["203", " SEM MATCH ", "07/01/2026", "08/01/2026"])
            ws["B2"].fill = PatternFill("solid", fgColor="FFFF00")
            ws["C2"].number_format = "dd/mm/yyyy"
            ws.column_dimensions["A"].width = 18
            wb.save(planilha_2)

            result = executar_match(str(planilha_1), str(planilha_2), 65)
            self.assertTrue(result["ok"], result.get("erro"))

            write_excel(
                result["nomes_finais"],
                result["statuses"],
                output,
                result["scores"],
                template_path=result["template_path"],
                name_rows=result["name_rows"],
            )

            original = load_workbook(planilha_2)
            exported = load_workbook(output)

        original_ws = original.active
        exported_ws = exported.active
        self.assertEqual(exported_ws.title, "Template")
        self.assertEqual(exported_ws.max_row, original_ws.max_row)
        self.assertEqual(exported_ws.max_column, original_ws.max_column)
        self.assertEqual(exported_ws[1][0].value, "room number")
        self.assertEqual(exported_ws["B1"].value, "name")
        self.assertEqual(exported_ws["A2"].value, "201")
        self.assertEqual(exported_ws["C2"].value, "03/01/2026")
        self.assertEqual(exported_ws["D4"].value, "08/01/2026")
        self.assertEqual(exported_ws["B2"].value, "Bruno Costa")
        self.assertEqual(exported_ws["B3"].value, "Ana Maria Oliveira")
        self.assertEqual(exported_ws["B4"].value, " SEM MATCH ")
        self.assertEqual(exported_ws["B2"].fill.fgColor.rgb, original_ws["B2"].fill.fgColor.rgb)
        self.assertEqual(exported_ws.column_dimensions["A"].width, original_ws.column_dimensions["A"].width)

    def test_reference_name_is_not_reused_for_later_weak_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "planilha_1.xlsx"
            planilha_2 = base / "planilha_2.xlsx"

            _write_names(planilha_1, ["Helton Machado Kraus"])
            _write_names(planilha_2, ["HELTON KRAUS", "LAIS MACHADO KLEIN"])

            result = executar_match(str(planilha_1), str(planilha_2), 65)

        self.assertTrue(result["ok"], result.get("erro"))
        self.assertEqual(
            result["nomes_finais"],
            ["Helton Machado Kraus", "LAIS MACHADO KLEIN"],
        )
        self.assertEqual(result["statuses"], ["Match encontrado", "Nao encontrado"])
        self.assertEqual(result["kpis"], {"total": 2, "match": 1, "nomatch": 1, "empty": 0})

    def test_reference_name_is_assigned_to_highest_scoring_row(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "planilha_1.xlsx"
            planilha_2 = base / "planilha_2.xlsx"

            _write_names(planilha_1, ["Ana Maria Oliveira"])
            _write_names(planilha_2, ["ANA OLIVEIRA", "ANA MARIA OLIVEIRA"])

            result = executar_match(str(planilha_1), str(planilha_2), 65)

        self.assertTrue(result["ok"], result.get("erro"))
        self.assertEqual(
            result["nomes_finais"],
            ["ANA OLIVEIRA", "Ana Maria Oliveira"],
        )
        self.assertEqual(result["statuses"], ["Nao encontrado", "Match encontrado"])
        self.assertEqual(result["scores"], [95.2, 100.0])

    def test_duplicate_reference_names_can_each_be_matched_once(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "planilha_1.xlsx"
            planilha_2 = base / "planilha_2.xlsx"

            _write_names(planilha_1, ["Helton Machado Kraus", "Helton Machado Kraus"])
            _write_names(planilha_2, ["HELTON KRAUS", "HELTON KRAUS"])

            result = executar_match(str(planilha_1), str(planilha_2), 65)

        self.assertTrue(result["ok"], result.get("erro"))
        self.assertEqual(
            result["nomes_finais"],
            ["Helton Machado Kraus", "Helton Machado Kraus"],
        )
        self.assertEqual(result["statuses"], ["Match encontrado", "Match encontrado"])
        self.assertEqual(result["kpis"], {"total": 2, "match": 2, "nomatch": 0, "empty": 0})


if __name__ == "__main__":
    unittest.main()
