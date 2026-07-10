import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.bate_rooming import Status, normalize_dates, parse_date_raw, processar_arquivos
from core.bate_rooming_export import write_excel as write_bate_excel
from core.match_nomes import executar_match


def _write_rows(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


class BusinessRuleCharacterizationTests(unittest.TestCase):
    def test_dates_accept_supported_formats_and_normalize_outlier_year(self):
        self.assertEqual(parse_date_raw("31/12/2026"), date(2026, 12, 31))
        self.assertEqual(parse_date_raw("2026-12-31"), date(2026, 12, 31))
        self.assertIsNone(parse_date_raw("data invalida"))

        normalized = normalize_dates(
            [date(2026, 1, 10), date(2026, 2, 10), date(2036, 3, 10)]
        )

        self.assertEqual(normalized, [date(2026, 1, 10), date(2026, 2, 10), date(2026, 3, 10)])

    def test_bate_rooming_marks_duplicates_and_placeholders_as_repeated(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(
                sistema,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["101", "Nome Duplicado", "01/01/2026", "02/01/2026"],
                    ["102", "Nome Duplicado", "01/01/2026", "02/01/2026"],
                    ["103", "A nomear", "01/01/2026", "02/01/2026"],
                ],
            )
            _write_rows(
                hotel,
                [["quarto", "nome", "check in", "check out"], ["201", "Outro Nome", "01/01/2026", "02/01/2026"]],
            )

            results, warnings, kpis = processar_arquivos(str(sistema), str(hotel))

        repeated = [row for row in results if row["s_geral"] == Status.REPETIDO]
        self.assertEqual(len(repeated), 3)
        self.assertEqual(kpis["dups"], 3)
        self.assertTrue(any("repetidos/placeholders" in warning for warning in warnings))

    def test_bate_rooming_uses_match_normalization_for_duplicates(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(
                sistema,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["101", "Nome Duplicado", "01/01/2026", "02/01/2026"],
                    ["102", "Nome-Duplicado", "01/01/2026", "02/01/2026"],
                ],
            )
            _write_rows(
                hotel,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["201", "NOME DUPLICADO", "01/01/2026", "02/01/2026"],
                ],
            )

            results, warnings, kpis = processar_arquivos(str(sistema), str(hotel))

        self.assertEqual(len(results), 3)
        self.assertTrue(all(row["s_geral"] == Status.REPETIDO for row in results))
        self.assertEqual(kpis["dups"], 3)
        self.assertTrue(any("nome-duplicado" in warning.lower() for warning in warnings))

    def test_bate_rooming_uses_highest_score_with_same_identity(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(
                sistema,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["101", "Ana Patricia Silva", "01/01/2026", "02/01/2026"],
                    ["102", "Ana Valentina Silva", "01/01/2026", "02/01/2026"],
                ],
            )
            _write_rows(
                hotel,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["202", "Ana Valentina da Silva", "01/01/2026", "02/01/2026"],
                ],
            )

            results, _, kpis = processar_arquivos(str(sistema), str(hotel))

        matched = [row for row in results if not row["no_match"]]
        unmatched = [row for row in results if row["no_match"]]
        self.assertEqual(kpis["matched"], 1)
        self.assertEqual(matched[0]["nome"], "Ana Valentina Silva")
        self.assertEqual(matched[0]["match_audit"]["stage"], "MESMA_IDENTIDADE_MAIOR_SCORE")
        self.assertGreaterEqual(matched[0]["match_audit"]["score"], 99.0)
        self.assertEqual(unmatched[0]["nome"], "Ana Patricia Silva")

    def test_bate_rooming_blocks_low_confidence_cross_identity_match(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(sistema, [["quarto", "nome"], ["101", "Ana Paula"]])
            _write_rows(hotel, [["quarto", "nome"], ["101", "Ana Clara"]])

            results, _, kpis = processar_arquivos(str(sistema), str(hotel))

        self.assertEqual(kpis["nomap"], 2)
        self.assertTrue(all(row["no_match"] for row in results))
        hotel_row = next(row for row in results if row["fonte"] == "HOTEL")
        self.assertGreater(hotel_row["match_audit"]["candidates_blocked"], 0)
        self.assertIn("falso positivo", hotel_row["match_audit"]["reason"])

    def test_bate_log_records_reference_conflict_and_winner(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(sistema, [["quarto", "nome"], ["101", "Maria Silva"]])
            _write_rows(
                hotel,
                [["quarto", "nome"], ["201", "Maria Silv"], ["202", "Maria Silvaa"]],
            )

            results, _, kpis = processar_arquivos(str(sistema), str(hotel))

        self.assertEqual(kpis["matched"], 1)
        unmatched = next(row for row in results if row["nome"] == "Maria Silv")
        conflict = unmatched["match_audit"]["conflicts"][0]
        self.assertEqual(conflict["nome_referencia"], "Maria Silva")
        self.assertEqual(conflict["winner_nome"], "Maria Silvaa")
        self.assertGreater(conflict["winner_score"], conflict["score"])

    def test_bate_export_includes_matching_log(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            output = base / "resultado.xlsx"
            _write_rows(sistema, [["quarto", "nome"], ["101", "Maria Silva"]])
            _write_rows(hotel, [["quarto", "nome"], ["101", "Maria Souza"]])

            results, _, _ = processar_arquivos(str(sistema), str(hotel))
            write_bate_excel(results, output)
            wb = load_workbook(output, data_only=True)
            sheetnames = wb.sheetnames
            log = wb["LOG"]
            headers = [cell.value for cell in log[2]]
            log_values = list(log.iter_rows(min_row=3, values_only=True))
            wb.close()

        self.assertIn("LOG", sheetnames)
        self.assertIn("DECISAO", headers)
        self.assertIn("TOP CANDIDATOS", headers)
        self.assertIn("BLOQUEADO_REGRA", " ".join(str(value or "") for row in log_values for value in row))

    def test_bate_rooming_uses_match_normalization_for_placeholders(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(
                sistema,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["101", "A-Nomear", "01/01/2026", "02/01/2026"],
                ],
            )
            _write_rows(
                hotel,
                [["quarto", "nome", "check in", "check out"], ["201", "Outro Nome", "01/01/2026", "02/01/2026"]],
            )

            results, warnings, kpis = processar_arquivos(str(sistema), str(hotel))

        repeated = [row for row in results if row["s_geral"] == Status.REPETIDO]
        self.assertEqual(len(repeated), 1)
        self.assertEqual(kpis["dups"], 1)
        self.assertTrue(any("a-nomear" in warning.lower() for warning in warnings))

    def test_bate_rooming_warns_when_optional_columns_are_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(sistema, [["nome"], ["Maria Silva"]])
            _write_rows(hotel, [["nome"], ["Maria Silva"]])

            results, warnings, kpis = processar_arquivos(str(sistema), str(hotel))

        self.assertEqual(len(results), 1)
        self.assertEqual(kpis["matched"], 1)
        self.assertEqual(kpis["div"], 1)
        self.assertEqual(len(warnings), 6)
        self.assertTrue(all("tratados como ausentes" in warning for warning in warnings))

    def test_ignore_room_preserves_date_rules_and_hides_room_columns_on_export(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            output = base / "resultado.xlsx"
            _write_rows(
                sistema,
                [["quarto", "nome", "check in", "check out"], ["101", "Maria Silva", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                hotel,
                [["quarto", "nome", "check in", "check out"], ["999", "Maria Silva", "01/01/2026", "03/01/2026"]],
            )

            results, _, kpis = processar_arquivos(str(sistema), str(hotel), ignorar_quarto=True)
            write_bate_excel(results, output)
            wb = load_workbook(output, data_only=True)
            headers = [cell.value for cell in wb["RESULTADO COMPLETO"][2]]
            wb.close()

        self.assertEqual(results[0]["s_quarto"], Status.IGNORADO)
        self.assertEqual(results[0]["s_geral"], Status.OK)
        self.assertEqual(kpis["ok"], 1)
        self.assertNotIn("QUARTO (PLAN. 1)", headers)
        self.assertNotIn("QUARTO (PLAN. 2)", headers)
        self.assertNotIn("STATUS QUARTO", headers)

    def test_match_nomes_preserves_unmatched_and_empty_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            reference = base / "reference.xlsx"
            target = base / "target.xlsx"
            _write_rows(reference, [["nome"], ["Maria Silva"]])
            _write_rows(target, [["nome"], ["Pessoa Sem Par"], [None]])

            result = executar_match(str(reference), str(target), 95)

        self.assertTrue(result["ok"], result.get("erro"))
        self.assertEqual(result["nomes_finais"], ["Pessoa Sem Par", ""])
        self.assertEqual(result["statuses"], ["Nao encontrado", "NOME VAZIO"])
        self.assertEqual(result["kpis"], {"total": 2, "match": 0, "nomatch": 1, "empty": 1})


if __name__ == "__main__":
    unittest.main()
