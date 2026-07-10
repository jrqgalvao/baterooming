import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.bate_rooming import Status, processar_arquivos
from core.bate_rooming_export import write_excel as write_bate_excel
from core.matching import ProcessingCancelled
from core.match_nomes import executar_match
from core.match_nomes_export import write_excel as write_match_excel


def _write_rows(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


class CoreBehaviorTests(unittest.TestCase):
    def test_bate_rooming_applies_integrated_name_match_without_audit_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            output = base / "resultado.xlsx"
            _write_rows(
                sistema,
                [["quarto", "nome", "check in", "check out"], ["101", "Helton Machado Kraus", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                hotel,
                [["quarto", "nome", "check in", "check out"], ["101", "HELTON KRAUS", "01/01/2026", "03/01/2026"]],
            )

            results, warnings, kpis = processar_arquivos(str(sistema), str(hotel))
            write_bate_excel(results, output)
            wb = load_workbook(output, data_only=True)
            headers = [cell.value for cell in wb["RESULTADO COMPLETO"][2]]
            wb.close()

        self.assertEqual(warnings, [])
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["nome"], "Helton Machado Kraus")
        self.assertFalse(results[0]["no_match"])
        self.assertEqual(results[0]["s_geral"], Status.OK)
        self.assertEqual(kpis["ok"], 1)
        self.assertNotIn("SCORE MATCH NOME", headers)
        self.assertNotIn("STATUS MATCH NOME", headers)

    def test_bate_rooming_does_not_reuse_reference_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(
                sistema,
                [["quarto", "nome", "check in", "check out"], ["101", "Helton Machado Kraus", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                hotel,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["101", "HELTON KRAUS", "01/01/2026", "03/01/2026"],
                    ["202", "LAIS MACHADO KLEIN", "01/01/2026", "03/01/2026"],
                ],
            )

            results, _, kpis = processar_arquivos(str(sistema), str(hotel))

        self.assertEqual(len([r for r in results if not r["no_match"]]), 1)
        self.assertEqual(len([r for r in results if r["no_match"]]), 1)
        self.assertEqual(kpis["nomap"], 1)

    def test_bate_rooming_prioritizes_same_first_and_last_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(
                sistema,
                [["quarto", "nome", "check in", "check out"], ["202", "Juliana Fernandes", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                hotel,
                [
                    ["quarto", "nome", "check in", "check out"],
                    ["101", "JULIANA FERNANDES COSTA", "01/01/2026", "03/01/2026"],
                    ["202", "JULIANA COSTA FERNANDES", "01/01/2026", "03/01/2026"],
                ],
            )

            results, _, kpis = processar_arquivos(str(sistema), str(hotel))

        matched = [row for row in results if not row["no_match"]]
        self.assertEqual(len(matched), 1)
        self.assertEqual(matched[0]["nome"], "Juliana Fernandes")
        self.assertEqual(matched[0]["q_hotel"], "202")
        self.assertEqual(kpis["nomap"], 1)

    def test_match_nomes_independent_flow_keeps_statuses_scores_and_template_export(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "planilha_1.xlsx"
            planilha_2 = base / "planilha_2.xlsx"
            output = base / "corrigido.xlsx"
            _write_rows(
                planilha_1,
                [["room number", "name", "check in date", "check out date"], ["101", "Helton Machado Kraus", "01/01/2026", "02/01/2026"]],
            )
            _write_rows(
                planilha_2,
                [["room number", "name", "check in date", "check out date"], ["201", "HELTON KRAUS", "03/01/2026", "04/01/2026"]],
            )

            result = executar_match(str(planilha_1), str(planilha_2), 65)
            write_match_excel(
                result["nomes_finais"],
                result["statuses"],
                output,
                result["scores"],
                template_path=result["template_path"],
                name_rows=result["name_rows"],
            )
            wb = load_workbook(output, data_only=True)
            corrected_name = wb.active["B2"].value
            wb.close()

        self.assertTrue(result["ok"], result.get("erro"))
        self.assertEqual(result["nomes_finais"], ["Helton Machado Kraus"])
        self.assertEqual(result["statuses"], ["Match encontrado"])
        self.assertGreaterEqual(result["scores"][0], 65)
        self.assertEqual(corrected_name, "Helton Machado Kraus")

    def test_progress_callbacks_preserve_completed_results(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            sistema = base / "sistema.xlsx"
            hotel = base / "hotel.xlsx"
            _write_rows(
                sistema,
                [["quarto", "nome", "check in", "check out"], ["101", "Maria Silva", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                hotel,
                [["quarto", "nome", "check in", "check out"], ["101", "MARIA SILVA", "01/01/2026", "03/01/2026"]],
            )

            expected_bate = processar_arquivos(str(sistema), str(hotel))
            bate_progress = []
            actual_bate = processar_arquivos(
                str(sistema),
                str(hotel),
                progress=lambda percent, message: bate_progress.append((percent, message)),
            )
            expected_match = executar_match(str(sistema), str(hotel), 65)
            match_progress = []
            actual_match = executar_match(
                str(sistema),
                str(hotel),
                65,
                progress=lambda percent, message: match_progress.append((percent, message)),
            )

        self.assertEqual(actual_bate, expected_bate)
        self.assertEqual(actual_match, expected_match)
        self.assertTrue(bate_progress)
        self.assertTrue(match_progress)

    def test_processing_can_cancel_before_opening_files(self):
        with self.assertRaises(ProcessingCancelled):
            processar_arquivos(
                "arquivo_1_inexistente.xlsx",
                "arquivo_2_inexistente.xlsx",
                should_cancel=lambda: True,
            )

        with self.assertRaises(ProcessingCancelled):
            executar_match(
                "arquivo_1_inexistente.xlsx",
                "arquivo_2_inexistente.xlsx",
                65,
                should_cancel=lambda: True,
            )


if __name__ == "__main__":
    unittest.main()
