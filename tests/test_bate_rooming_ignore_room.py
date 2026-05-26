import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bate_rooming_core import Status, processar_arquivos, write_excel


def _write_rows(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


class BateRoomingIgnoreRoomTests(unittest.TestCase):
    def test_default_mode_keeps_room_divergence(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "sistema.xlsx"
            planilha_2 = base / "hotel.xlsx"
            _write_rows(
                planilha_1,
                [["quarto", "nome", "check in", "check out"], ["101", "Ana Silva", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                planilha_2,
                [["quarto", "nome", "check in", "check out"], ["909", "Ana Silva", "01/01/2026", "03/01/2026"]],
            )

            results, warnings, kpis = processar_arquivos(str(planilha_1), str(planilha_2))

        self.assertEqual(warnings, [])
        self.assertEqual(results[0]["s_quarto"], Status.DIVERGENTE)
        self.assertEqual(kpis["q_div"], 1)
        self.assertEqual(kpis["div"], 1)

    def test_ignore_room_mode_excludes_room_from_divergences(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "sistema.xlsx"
            planilha_2 = base / "hotel.xlsx"
            _write_rows(
                planilha_1,
                [["quarto", "nome", "check in", "check out"], ["101", "Ana Silva", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                planilha_2,
                [["quarto", "nome", "check in", "check out"], ["909", "Ana Silva", "01/01/2026", "03/01/2026"]],
            )

            results, warnings, kpis = processar_arquivos(
                str(planilha_1),
                str(planilha_2),
                ignorar_quarto=True,
            )

        self.assertEqual(warnings, [])
        self.assertEqual(results[0]["s_quarto"], Status.IGNORADO)
        self.assertTrue(results[0]["room_check_ignored"])
        self.assertEqual(results[0]["s_geral"], Status.OK)
        self.assertEqual(kpis["q_div"], 0)
        self.assertEqual(kpis["div"], 0)
        self.assertEqual(kpis["ok"], 1)

    def test_ignore_room_excel_output_hides_room_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "sistema.xlsx"
            planilha_2 = base / "hotel.xlsx"
            export_path = base / "resultado.xlsx"
            _write_rows(
                planilha_1,
                [["quarto", "nome", "check in", "check out"], ["101", "Ana Silva", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                planilha_2,
                [["quarto", "nome", "check in", "check out"], ["909", "Ana Silva", "01/01/2026", "03/01/2026"]],
            )

            results, _, _ = processar_arquivos(str(planilha_1), str(planilha_2), ignorar_quarto=True)
            write_excel(results, export_path)
            wb = load_workbook(export_path)
            headers = [cell.value for cell in wb["RESULTADO COMPLETO"][2] if cell.value]
            summary_labels = [
                cell.value
                for row in wb["RESUMO"].iter_rows()
                for cell in row
                if isinstance(cell.value, str)
            ]
            wb.close()

        self.assertNotIn("QUARTO (PLAN. 1)", headers)
        self.assertNotIn("QUARTO (PLAN. 2)", headers)
        self.assertNotIn("STATUS QUARTO", headers)
        self.assertIn("STATUS GERAL", headers)
        self.assertNotIn("Divergência — Quarto", summary_labels)

    def test_ignore_room_excel_status_general_omits_room_label_when_other_fields_differ(self):
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            planilha_1 = base / "sistema.xlsx"
            planilha_2 = base / "hotel.xlsx"
            export_path = base / "resultado.xlsx"
            _write_rows(
                planilha_1,
                [["quarto", "nome", "check in", "check out"], ["101", "Ana Silva", "01/01/2026", "03/01/2026"]],
            )
            _write_rows(
                planilha_2,
                [["quarto", "nome", "check in", "check out"], ["909", "Ana Silva", "02/01/2026", "03/01/2026"]],
            )

            results, _, _ = processar_arquivos(str(planilha_1), str(planilha_2), ignorar_quarto=True)
            write_excel(results, export_path)
            wb = load_workbook(export_path)
            headers = [cell.value for cell in wb["RESULTADO COMPLETO"][2] if cell.value]
            status_col = headers.index("STATUS GERAL") + 1
            status_geral = wb["RESULTADO COMPLETO"].cell(3, status_col).value
            wb.close()

        self.assertEqual(results[0]["s_quarto"], Status.IGNORADO)
        self.assertEqual(results[0]["s_ci"], Status.CHECK_IN)
        self.assertEqual(status_geral, Status.CHECK_IN)
        self.assertNotIn("QUARTO", str(status_geral))


if __name__ == "__main__":
    unittest.main()
