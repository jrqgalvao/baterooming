"""
match_nomes_core.py — Match de Nomes v3 (Your Company)
Lógica de fuzzy matching com pipeline multi-etapa e exportação Excel.

Pipeline de matching (em ordem de prioridade):
  1. Exato normalizado          — score 100, custo zero
  2. Token set ratio (rapidfuzz) — lida com palavras fora de ordem
  3. Token sort ratio            — lida com inversão de nome/sobrenome
  4. Partial ratio               — lida com nomes abreviados/incompletos
  5. Weighted best               — combinação ponderada dos scores acima

Sem dependências de UI.

Dependências:
    pip install pandas rapidfuzz openpyxl
"""

import unicodedata
import re
from functools import lru_cache
from pathlib import Path
from datetime import datetime

import pandas as pd
from rapidfuzz import fuzz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


MATCH_FOUND = "Match encontrado"
NOT_FOUND = "Nao encontrado"
EMPTY_NAME = "NOME VAZIO"

_RE_NON_ALNUM_SPACE = re.compile(r"[^a-z0-9\s]")
_RE_SPACES = re.compile(r"\s+")


@lru_cache(maxsize=8192)
def _normalizar_texto(nome: str) -> str:
    nome = unicodedata.normalize("NFD", nome)
    nome = "".join(c for c in nome if unicodedata.category(c) != "Mn")
    nome = nome.lower()
    nome = _RE_NON_ALNUM_SPACE.sub(" ", nome)
    nome = _RE_SPACES.sub(" ", nome).strip()
    return nome


# ── Normalização ───────────────────────────────────────────────
def normalizar(nome: str) -> str:
    """Remove acentos, lowercase, colapsa espacos."""
    if not isinstance(nome, str) or nome.strip() == "":
        return ""
    return _normalizar_texto(nome)


# Palavras que reduzem sinal quando isoladas (partículas de nome)
_STOPWORDS = {"de", "da", "do", "das", "dos", "e", "van", "von", "del", "di", "la", "le"}
_NOME_HEADER_ALIASES = {"nome", "name", "hospede", "guest", "cliente", "passageiro"}
_DEFAULT_NAME_COLUMN = 2


def tokens_significativos(nome_norm: str) -> list[str]:
    return [t for t in nome_norm.split() if t not in _STOPWORDS and len(t) > 1]


def _header_slug(value) -> str:
    return normalizar(str(value or "")).replace(" ", "").replace("_", "")


def _find_name_header(ws) -> "tuple[int, int] | None":
    max_scan_rows = min(5, ws.max_row or 0)
    max_scan_cols = ws.max_column or 0
    for row_idx in range(1, max_scan_rows + 1):
        for col_idx in range(1, max_scan_cols + 1):
            if _header_slug(ws.cell(row=row_idx, column=col_idx).value) in _NOME_HEADER_ALIASES:
                return col_idx, row_idx
    return None


def _detect_name_layout(ws) -> tuple[int, int]:
    header = _find_name_header(ws)
    if header is not None:
        name_col, header_row = header
        return name_col, header_row + 1

    if (ws.max_column or 0) >= 4:
        name_col = _DEFAULT_NAME_COLUMN
    else:
        name_col = 1

    return name_col, 1


def _read_nomes(path: str) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws.max_row == 0:
            return pd.DataFrame({"nome": [], "row_number": [], "name_column": []})

        name_col, start_row = _detect_name_layout(ws)
        max_row = ws.max_row or 0
        nomes = []
        row_numbers = []
        for row_idx, row in enumerate(
            ws.iter_rows(
                min_row=start_row,
                max_row=max_row,
                min_col=name_col,
                max_col=name_col,
                values_only=True,
            ),
            start=start_row,
        ):
            nomes.append(row[0] if row else None)
            row_numbers.append(row_idx)
        return pd.DataFrame({"nome": nomes, "row_number": row_numbers, "name_column": [name_col] * len(nomes)})
    finally:
        wb.close()


def _nome_limpo(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text


# ── Scoring multi-etapa ────────────────────────────────────────
def _score_par(
    query: str,
    candidate: str,
    tq: "set | None" = None,
    tc: "set | None" = None,
) -> float:
    """
    Retorna score 0-100 entre query e candidate (ambos já normalizados).

    Combina quatro métricas com pesos:
      - token_set_ratio   (0.40): robusto a palavras fora de ordem
      - token_sort_ratio  (0.30): robusto a inversão nome/sobrenome
      - partial_ratio     (0.20): robusto a nomes abreviados
      - ratio             (0.10): similaridade global de string

    Bônus de +5 se todos os tokens significativos do candidato menor
    aparecem no maior (match por subconjunto de tokens).

    O parâmetro opcional `tq` permite passar os tokens do query
    pré-calculados externamente, evitando recálculo por candidato.
    """
    if not query or not candidate:
        return 0.0

    s_set  = fuzz.token_set_ratio(query, candidate)
    s_sort = fuzz.token_sort_ratio(query, candidate)
    s_part = fuzz.partial_ratio(query, candidate)
    s_full = fuzz.ratio(query, candidate)

    score = (
        0.40 * s_set  +
        0.30 * s_sort +
        0.20 * s_part +
        0.10 * s_full
    )

    # Bônus: subconjunto de tokens
    if tq is None:
        tq = set(tokens_significativos(query))
    if tc is None:
        tc = set(tokens_significativos(candidate))
    if tq and tc:
        menor, maior = (tq, tc) if len(tq) <= len(tc) else (tc, tq)
        if menor and menor.issubset(maior):
            score = min(100.0, score + 5.0)

    return score


def _melhor_match_com_indice(
    query_norm: str,
    lista_norm: list[str],
    lista_orig: list[str],
    threshold: float,
    exact_lookup: "dict[str, list[int]] | None" = None,
    lista_tokens: "list[set[str]] | None" = None,
    indices_usados: "set[int] | None" = None,
) -> tuple[int, str, float, str]:
    """
    Retorna o indice escolhido junto do melhor nome encontrado.
    Isso permite impedir que uma linha da Planilha 1 seja reutilizada.
    """
    indices_usados = indices_usados or set()

    if not query_norm:
        return (-1, "", 0.0, EMPTY_NAME)

    # Passo 1: match exato normalizado (custo zero)
    if exact_lookup is not None:
        for exact_idx in exact_lookup.get(query_norm, []):
            if exact_idx not in indices_usados:
                return (exact_idx, lista_orig[exact_idx], 100.0, MATCH_FOUND)
    else:
        for i, cand in enumerate(lista_norm):
            if i not in indices_usados and cand == query_norm:
                return (i, lista_orig[i], 100.0, MATCH_FOUND)

    # Pré-calcula os tokens do query uma única vez para toda a iteração
    tq = set(tokens_significativos(query_norm))

    # Passo 2: busca exaustiva com score ponderado
    if not lista_norm or len(indices_usados) >= len(lista_norm):
        return (-1, "", 0.0, NOT_FOUND)

    best_score = -1.0
    best_idx   = -1
    for i, cand in enumerate(lista_norm):
        if i in indices_usados:
            continue
        tc = lista_tokens[i] if lista_tokens is not None else None
        s = _score_par(query_norm, cand, tq, tc)
        if s > best_score:
            best_score = s
            best_idx   = i

    if best_idx >= 0 and best_score >= threshold:
        return (best_idx, lista_orig[best_idx], best_score, MATCH_FOUND)

    return (-1, "", max(best_score, 0.0), NOT_FOUND)


def _atribuir_melhores_matches(
    nomes_b_originais: list[str],
    nomes_b_normalizados: list[str],
    lista_a_original: list[str],
    lista_a_normalizada: list[str],
    lista_a_tokens: list[set[str]],
    threshold: float,
) -> tuple[list[str], list[str], list[float]]:
    """
    Atribui cada linha da Planilha 1 a no maximo uma linha da Planilha 2.
    Duplicatas reais na Planilha 1 continuam disponiveis como linhas separadas.
    """
    nomes_finais = list(nomes_b_originais)
    statuses = [EMPTY_NAME if not nome_norm else NOT_FOUND for nome_norm in nomes_b_normalizados]
    scores = [0.0 for _ in nomes_b_originais]
    propostas: list[tuple[float, int, int]] = []

    for b_idx, nome_b_norm in enumerate(nomes_b_normalizados):
        if not nome_b_norm:
            continue

        tq = set(tokens_significativos(nome_b_norm))
        best_score = 0.0
        for a_idx, nome_a_norm in enumerate(lista_a_normalizada):
            score = 100.0 if nome_a_norm == nome_b_norm else _score_par(
                nome_b_norm,
                nome_a_norm,
                tq,
                lista_a_tokens[a_idx],
            )
            best_score = max(best_score, score)
            if score >= threshold:
                propostas.append((score, b_idx, a_idx))
        scores[b_idx] = round(best_score, 1)

    propostas.sort(key=lambda item: (-item[0], item[1], item[2]))

    linhas_b_atribuidas: set[int] = set()
    linhas_a_usadas: set[int] = set()
    for score, b_idx, a_idx in propostas:
        if b_idx in linhas_b_atribuidas or a_idx in linhas_a_usadas:
            continue
        nomes_finais[b_idx] = lista_a_original[a_idx]
        statuses[b_idx] = MATCH_FOUND
        scores[b_idx] = round(score, 1)
        linhas_b_atribuidas.add(b_idx)
        linhas_a_usadas.add(a_idx)

    return nomes_finais, statuses, scores


# ── Match principal ────────────────────────────────────────────
def executar_match(arquivo_a: str, arquivo_b: str, threshold: int) -> dict:
    """
    Executa o fuzzy match entre Lista A (referência) e Lista B (a corrigir).

    Regras de negócio:
      - A saída tem exatamente len(Lista B) linhas, na mesma ordem.
      - Linhas sem match suficiente mantêm o nome original e recebem
        status "Nao encontrado".
      - Linhas vazias recebem status "NOME VAZIO".

    Retorna:
        {
            "ok": True,
            "nomes_finais": [...],
            "statuses":     [...],
            "scores":       [...],   # float 0-100, para debug/auditoria
            "kpis": {"total": N, "match": N, "nomatch": N, "empty": N}
        }
        ou {"ok": False, "erro": "..."}
    """
    try:
        df_a = _read_nomes(arquivo_a)
        df_b = _read_nomes(arquivo_b)

        df_a["nome"] = df_a["nome"].astype(str).str.strip()
        df_b["nome"] = df_b["nome"].astype(str).str.strip()

        # Remove linhas totalmente vazias da lista de referência
        df_a = df_a[~df_a["nome"].isin(["", "nan", "None"])].reset_index(drop=True)

        lista_a_original    = df_a["nome"].tolist()
        lista_a_normalizada = [normalizar(n) for n in lista_a_original]
        lista_a_tokens      = [set(tokens_significativos(n)) for n in lista_a_normalizada]
        thr_float = float(threshold)
        nomes_b = [_nome_limpo(nome_b_raw) for nome_b_raw in df_b["nome"].tolist()]
        nomes_b_normalizados = [normalizar(nome_b_clean) for nome_b_clean in nomes_b]

        nomes_finais, statuses, scores = _atribuir_melhores_matches(
            nomes_b,
            nomes_b_normalizados,
            lista_a_original,
            lista_a_normalizada,
            lista_a_tokens,
            thr_float,
        )

        total = len(df_b)
        match_count = statuses.count(MATCH_FOUND)
        nomatch_count = statuses.count(NOT_FOUND)
        empty_count = statuses.count(EMPTY_NAME)

        kpis = {
            "total":   total,
            "match":   match_count,
            "nomatch": nomatch_count,
            "empty":   empty_count,
        }

        return {
            "ok":           True,
            "nomes_finais": nomes_finais,
            "statuses":     statuses,
            "scores":       scores,
            "kpis":         kpis,
            "template_path": str(arquivo_b),
            "name_rows":     df_b["row_number"].tolist(),
            "name_column":   int(df_b["name_column"].iloc[0]) if len(df_b) else _DEFAULT_NAME_COLUMN,
        }

    except Exception as exc:
        return {"ok": False, "erro": str(exc).strip() or "Não foi possível processar as planilhas."}


# ── Exportação Excel ───────────────────────────────────────────
_THIN = Border(
    left   = Side(style="thin", color="C8D8C8"),
    right  = Side(style="thin", color="C8D8C8"),
    top    = Side(style="thin", color="C8D8C8"),
    bottom = Side(style="thin", color="C8D8C8"),
)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

_STATUS_STYLES: dict[str, dict] = {
    "Match encontrado": {"fill": "E6F4E7", "font_color": "1E6B22"},
    "Nao encontrado":   {"fill": "FFF3E0", "font_color": "BF360C"},
    "NOME VAZIO":       {"fill": "F4F6F4", "font_color": "7A9A7A"},
}


def write_excel(
    nomes_finais : list[str],
    statuses     : list[str],
    path         : Path,
    scores       : "list[float] | None" = None,
    template_path: "str | Path | None" = None,
    name_rows    : "list[int] | None" = None,
    name_column  : int = _DEFAULT_NAME_COLUMN,
) -> None:
    """
    Quando `template_path` e `name_rows` sao fornecidos, salva uma copia da
    Planilha 2 e substitui somente os nomes encontrados na coluna de nomes.

    Sem template, mantem o export legado com colunas:
      nome_final | status | similaridade (%)
    """
    if template_path is not None and name_rows is not None:
        wb = load_workbook(str(template_path))
        ws = wb.active
        for row_idx, nome_final, status in zip(name_rows, nomes_finais, statuses):
            if status == MATCH_FOUND:
                ws.cell(row=int(row_idx), column=int(name_column)).value = nome_final
        wb.save(str(path))
        return

    include_score = scores is not None and len(scores) == len(nomes_finais)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    n_cols = 3 if include_score else 2
    headers = ["nome_final", "status"]
    if include_score:
        headers.append("similaridade_%")
    ws.append(headers)
    if include_score:
        for row in zip(nomes_finais, statuses, scores):
            ws.append(row)
    else:
        for row in zip(nomes_finais, statuses):
            ws.append(row)

    # ── Cabeçalho ──
    HDR_FILL  = _fill("141E14")
    HDR_FONT  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center")

    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = HDR_ALIGN
        c.border    = _THIN
    ws.row_dimensions[1].height = 26

    # ── Dados ──
    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row_idx, 2).value or "")
        style  = _STATUS_STYLES.get(status, {"fill": "FFFFFF", "font_color": "162016"})

        for col in range(1, n_cols + 1):
            c = ws.cell(row_idx, col)
            is_score_col = include_score and col == 3
            c.font = Font(
                name       = "Segoe UI",
                color      = style["font_color"],
                size       = 10,
                bold       = (col == 2),
                italic     = is_score_col,
            )
            c.fill      = _fill(style["fill"])
            c.alignment = Alignment(
                horizontal = "left"   if col == 1 else "center",
                vertical   = "center",
            )
            c.border = _THIN
        ws.row_dimensions[row_idx].height = 18

    # ── Larguras e freeze ──
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    if include_score:
        ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "A2"

    # ── Rodapé ──
    foot_row = ws.max_row + 2
    last_col_letter = "C" if include_score else "B"
    ws.merge_cells(f"A{foot_row}:{last_col_letter}{foot_row}")
    f = ws.cell(
        foot_row, 1,
        f"Your Company © 2026 — Match de Nomes v3 · Gerado em "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    f.font      = Font(italic=True, color="9AAA9A", size=8)
    f.alignment = Alignment(horizontal="right")

    wb.save(str(path))
